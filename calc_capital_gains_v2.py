"""Capital gains report generator for Futu annual statement Excel files (v2).

v2 与 v1 唯一差别: 持仓队列从 FIFO 改为**移动加权平均**(Moving Weighted Average)。
每只 (code, currency, side) 只维护一条 (qty, avg_cost, avg_fee_per_unit, first_open_date),
开仓时按数量加权更新 avg, 平仓时 qty 减少而 avg 保持不变, 直到清零。

ASS/EXR 注入权利金的逻辑不变(流水里 "买入开仓 @ 行权价" 仍会被改成
"买入开仓 @ (行权价 - premium/股)"), 调整后的价格会被纳入 MWA。

Scans `YYYY_年度账单_*.xlsx`, replays trades chronologically, and produces
`capital_gains_report_v2.html` + `.json`.
"""
from __future__ import annotations

import argparse
import glob
import html
import json
import re
from collections import defaultdict, deque
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Optional

# --------------------------------------------------------------------------- #
# Constants & config
# --------------------------------------------------------------------------- #

# --- 默认值 (可被 script_config.json 覆盖) ---

TARGET_ACCOUNTS: set[str] = set()
TARGET_ACCOUNT: str = ""
DEFAULT_OPTION_MULTIPLIER: dict[str, Decimal] = {"USD": Decimal("100"), "HKD": Decimal("500")}
OPTION_UNDERLYING_TO_STOCK: dict[str, str] = {}
TAX_RATE: Decimal = Decimal("0.20")
TAX_FX_USD_TO_CNY: dict[int, Decimal] = {}
TAX_HKD_PEG_DIVISOR: Decimal = Decimal("7.8")
SKIP_DIRECTIONS: set[str] = set()
SKIP_CATEGORIES: set[str] = set()
OUTPUT_HTML: str = "capital_gains_report_v2.html"
OUTPUT_JSON: str = "capital_gains_report_v2.json"

_CONFIG_FILE = "script_config.json"

_CONFIG_LOADED = False
_CONFIG_OVERRIDES: list[dict] = []
_CONFIG_TAX: dict = {}


def _auto_detect_accounts(target_dir: Path) -> list[str]:
    """从所有 xlsx 文件中扫描账户号码。"""
    accounts: set[str] = set()
    sheets_to_scan = ["证券-交易流水", "证券-持仓总览", "证券-资产进出", "证券-资金进出"]
    for p in target_dir.glob("*_年度账单_*.xlsx"):
        try:
            wb = openpyxl.load_workbook(str(p), data_only=True)
            for sheet_name in sheets_to_scan:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                header = None
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        header = list(row)
                        if "账户号码" not in header:
                            break
                        continue
                    if not header:
                        break
                    rec = dict(zip(header, row))
                    acct = str(rec.get("账户号码") or "").strip()
                    if acct:
                        accounts.add(acct)
        except Exception:
            continue
    return sorted(accounts)


def load_config(target_dir: Path) -> dict:
    """加载 script_config.json。若 target_accounts 未填,自动从 xlsx 扫描。"""
    global _CONFIG_LOADED
    config_path = target_dir / _CONFIG_FILE

    if not config_path.exists():
        raise SystemExit(
            f"错误: 找不到配置文件 {_CONFIG_FILE}\n"
            f"请在 {target_dir} 目录下创建 {_CONFIG_FILE}。\n"
            f"示例:\n"
            f'  {{\n'
            f'    "default_option_multiplier": {{"USD": "100", "HKD": "500"}},\n'
            f'    "underlying_to_stock": {{"CMB": "3968"}},\n'
            f'    "tax_rate": "0.20",\n'
            f'    "fx_rates": {{"2023": "7.0827"}}\n'
            f'  }}'
        )

    raw = json.loads(config_path.read_text(encoding="utf-8"))

    # target_accounts 未填时自动从 xlsx 扫描
    accounts = raw.get("target_accounts")
    if not accounts or not isinstance(accounts, list) or len(accounts) == 0:
        detected = _auto_detect_accounts(target_dir)
        if not detected:
            raise SystemExit(
                f"错误: {_CONFIG_FILE} 未配置 target_accounts, 且未在 {target_dir} 的 xlsx 文件中检测到账户号码。\n"
                f"请确保目录中存在 年度账单 xlsx 文件, 或在 {_CONFIG_FILE} 中手动添加:\n"
                f'  "target_accounts": ["1001209838769859"]'
            )
        raw["target_accounts"] = detected
        print(f"[自动检测] 从 xlsx 中发现 {len(detected)} 个账户: {', '.join(detected)}")

    return raw


def apply_config(raw: dict) -> None:
    """将配置 JSON 应用到模块级全局变量。"""
    global TARGET_ACCOUNTS, TARGET_ACCOUNT
    global DEFAULT_OPTION_MULTIPLIER, OPTION_UNDERLYING_TO_STOCK
    global TAX_RATE, TAX_FX_USD_TO_CNY, TAX_HKD_PEG_DIVISOR
    global SKIP_DIRECTIONS, SKIP_CATEGORIES
    global OUTPUT_HTML, OUTPUT_JSON
    global _CONFIG_LOADED, _CONFIG_OVERRIDES, _CONFIG_TAX

    # 账户
    TARGET_ACCOUNTS = set(raw["target_accounts"])
    TARGET_ACCOUNT = raw["target_accounts"][0]

    # 期权乘数
    if "default_option_multiplier" in raw:
        DEFAULT_OPTION_MULTIPLIER = {
            k: Decimal(str(v)) for k, v in raw["default_option_multiplier"].items()
        }

    # 港股期权底层代码映射
    if "underlying_to_stock" in raw:
        OPTION_UNDERLYING_TO_STOCK = dict(raw["underlying_to_stock"])

    # 税务配置
    if "tax_rate" in raw:
        TAX_RATE = Decimal(str(raw["tax_rate"]))
    if "fx_rates" in raw:
        TAX_FX_USD_TO_CNY = {int(k): Decimal(str(v)) for k, v in raw["fx_rates"].items()}
    if "hkd_peg_divisor" in raw:
        TAX_HKD_PEG_DIVISOR = Decimal(str(raw["hkd_peg_divisor"]))

    # 过滤
    if "skip_categories" in raw:
        SKIP_CATEGORIES = set(raw["skip_categories"])
    if "skip_directions" in raw:
        SKIP_DIRECTIONS = set(raw["skip_directions"])

    # 输出文件名
    if "output_html" in raw:
        OUTPUT_HTML = raw["output_html"]
    if "output_json" in raw:
        OUTPUT_JSON = raw["output_json"]

    # Overrides (从 script_config.json 解析,可被 cost_basis_overrides.json 覆盖)
    _CONFIG_OVERRIDES = _parse_overrides_from_dict(raw)
    _CONFIG_TAX = _parse_tax_config_from_dict(raw)

    _CONFIG_LOADED = True

# --------------------------------------------------------------------------- #
# Dataclasses
# --------------------------------------------------------------------------- #

@dataclass
class Trade:
    """Normalized row from 证券-交易流水."""
    ts: datetime               # 成交时间
    category: str              # 证券 / 期权 / 基金
    code: str                  # 代码名称
    market: str
    currency: str              # HKD / USD
    direction: str             # 买入开仓/卖出开仓/买入平仓/卖出平仓/强平/申购/赎回
    quantity: Decimal          # 含正负号
    price: Decimal             # 单股/单合约成交价
    fee_total: Decimal         # 总费用(已聚合)
    amount_change: Decimal     # 变动金额(扣费后现金流)
    multiplier: Decimal        # 乘数(从持仓总览查表 + 兜底)
    note: str = ""             # 备注(如"由 AAPL231215C190000 CALL 被指派调整价格")

@dataclass
class AssetMovement:
    """Normalized row from 证券-资产进出 (期权 EXP-NA / ASS-P/C / EXR)."""
    ts: datetime               # 日期(精度到日,作为排序键)
    code: str                  # 期权合约代码
    currency: str
    quantity: Decimal          # 正数
    note: str                  # 原备注: 'Opt EXP-NA-...' / 'Opt ASS-P-...' etc.
    event: str                 # 'EXP-NA' / 'ASS-P' / 'ASS-C' / 'EXR-P' / 'EXR-C'

@dataclass
class PositionLot:
    """One FIFO entry: 一次建仓的余量."""
    code: str
    currency: str
    open_date: date
    open_price: Decimal        # 单股/单合约
    multiplier: Decimal
    quantity: Decimal          # 始终为正; 多头/空头由队列归属决定
    fee_per_unit: Decimal      # 平均到每单位的建仓费用
    note: str = ""             # 备注 (例如 '种子(2024期初)' )

@dataclass
class RealizedPnL:
    """A FIFO-matched close producing realized P/L."""
    realized_year: int
    code: str
    asset_type: str            # '股票' / '期权'
    direction: str             # '多头' / '空头'
    quantity: Decimal
    open_date: date
    open_price: Decimal
    close_date: date
    close_price: Decimal
    open_fee: Decimal          # 摊销后的建仓费用
    close_fee: Decimal         # 摊销后的平仓费用
    multiplier: Decimal
    currency: str
    pnl: Decimal               # 已实现盈亏
    note: str = ""
    underlying: str = ""       # 底层标的(股票=自身代码, 期权=OptionMeta.underlying 或港股映射后)
    sub_category: str = ""     # 细分: '到期作废' / '主动平仓' / '被指派/行权' / '现货买卖' / '种子持仓'

@dataclass
class OptionEvent:
    """期权指派/行权事件追踪 (展示用,不入合计)."""
    code: str
    event_type: str            # '到期作废' / 'PUT被指派' / 'CALL被指派' / 'PUT行权' / 'CALL行权'
    event_date: date
    contracts: Decimal
    strike: Optional[Decimal]
    premium_net: Optional[Decimal]    # 卖方权利金净额(空头)
    cost_paid: Optional[Decimal]      # 买方支付成本(多头)
    transferred_to_stock: bool
    note: str = ""

# --------------------------------------------------------------------------- #
# Section markers (to be filled by later tasks)
# --------------------------------------------------------------------------- #
# --------------------------------------------------------------------------- #
# Option-code parser
# --------------------------------------------------------------------------- #

_OPTION_RE = re.compile(r"^([A-Z]{1,6})(\d{6})([PC])(\d{4,8})$")


@dataclass
class OptionMeta:
    underlying: str
    expiry: date
    opt_type: str          # 'P' or 'C'
    strike: Decimal

    @property
    def underlying_stock_code(self) -> str:
        """HK 期权底层与股票代码不一致时,从映射查表;不在映射则 raise KeyError 提醒补全。
        US 期权调用方应直接使用 .underlying,不调用此属性。
        """
        if self.underlying in OPTION_UNDERLYING_TO_STOCK:
            return OPTION_UNDERLYING_TO_STOCK[self.underlying]
        raise KeyError(
            f"underlying {self.underlying!r} not in OPTION_UNDERLYING_TO_STOCK; "
            f"please add mapping"
        )


def parse_option_code(code: str) -> OptionMeta:
    """Parse Futu option contract code into structured fields.

    Format: <UND><YYMMDD><P|C><STRIKE×1000>
    Strike scaling: divide raw digits by 1000.
    """
    m = _OPTION_RE.match(code.strip())
    if not m:
        raise ValueError(f"invalid option code: {code!r}")
    und, ymd, opt_type, strike_raw = m.groups()
    yy, mm, dd = int(ymd[:2]), int(ymd[2:4]), int(ymd[4:6])
    year = 2000 + yy
    expiry = date(year, mm, dd)
    strike = (Decimal(strike_raw) / Decimal(1000)).quantize(Decimal("0.001"))
    return OptionMeta(underlying=und, expiry=expiry, opt_type=opt_type, strike=strike)


# Option-code parser:    Task 2 (done)
# --------------------------------------------------------------------------- #
# PositionBook
# --------------------------------------------------------------------------- #


@dataclass
class AvgPosition:
    """Per-(code, currency, side) 移动加权平均状态。"""
    code: str
    currency: str
    qty: Decimal                     # 当前持仓余量(始终非负;side 决定多/空)
    avg_cost: Decimal                # 每单位移动加权平均成本(不含费用)
    avg_fee_per_unit: Decimal        # 每单位累计建仓费用(平均到每股/张)
    first_open_date: date            # 首次建仓日(平仓不刷新, 清零后重新建仓时重置)
    multiplier: Decimal              # 该标的乘数(开仓时记录;多笔不同乘数视为冲突)
    last_note: str = ""              # 最近一次建仓的 note(便于追溯, 不影响计算)

    # --- v1 PositionLot 字段别名(供 render_html / JSON 复用 v1 代码) ---
    @property
    def quantity(self) -> Decimal:
        return self.qty

    @property
    def open_price(self) -> Decimal:
        return self.avg_cost

    @property
    def open_date(self) -> date:
        return self.first_open_date

    @property
    def fee_per_unit(self) -> Decimal:
        return self.avg_fee_per_unit

    @property
    def note(self) -> str:
        return self.last_note


class PositionBook:
    """Per-(code, currency, side) 移动加权平均(MWA)持仓表。

    与 FIFO 不同, 每只 (code, currency, side) 只维护一条 AvgPosition:
      - open(qty, price, fee):  按数量加权更新 avg_cost / avg_fee_per_unit
      - close(qty):             qty 减少, avg 保持不变(直到清零)
    清零后再次开仓将重置 first_open_date 与 avg。
    """

    def __init__(self) -> None:
        self._long: dict[tuple[str, str], AvgPosition] = {}
        self._short: dict[tuple[str, str], AvgPosition] = {}

    def _table(self, side: str) -> dict:
        if side == "long":
            return self._long
        if side == "short":
            return self._short
        raise ValueError(f"unknown side: {side!r}")

    def open(self, code: str, currency: str, side: str,
             qty: Decimal, price: Decimal, fee: Decimal,
             open_date: date, multiplier: Decimal,
             note: str = "") -> None:
        """累计开仓: 移动加权平均更新 avg_cost 和 avg_fee_per_unit。"""
        if qty <= 0:
            raise ValueError(f"open qty must be positive: {qty}")
        tab = self._table(side)
        key = (code, currency)
        pos = tab.get(key)
        if pos is None or pos.qty == 0:
            tab[key] = AvgPosition(
                code=code, currency=currency, qty=qty,
                avg_cost=price,
                avg_fee_per_unit=(fee / qty),
                first_open_date=open_date,
                multiplier=multiplier,
                last_note=note,
            )
            return
        # 加权累计
        new_qty = pos.qty + qty
        new_avg_cost = (pos.avg_cost * pos.qty + price * qty) / new_qty
        new_avg_fee = (pos.avg_fee_per_unit * pos.qty + fee) / new_qty
        pos.qty = new_qty
        pos.avg_cost = new_avg_cost
        pos.avg_fee_per_unit = new_avg_fee
        pos.last_note = note or pos.last_note
        # multiplier 不一致时以最新为准并不报错(对真实账单几乎不会出现混乱)
        if multiplier != pos.multiplier:
            pos.multiplier = multiplier

    def close(self, code: str, currency: str, side: str,
              qty: Decimal) -> tuple[Decimal, Decimal, Decimal, date, str]:
        """平仓 qty 单位, 返回 (avg_cost, amortized_open_fee, multiplier, first_open_date, last_note)。
        avg_cost 和 avg_fee_per_unit 不变;qty 减少;归零后等下次 open 重置。
        """
        if qty <= 0:
            raise ValueError(f"close qty must be positive: {qty}")
        tab = self._table(side)
        key = (code, currency)
        pos = tab.get(key)
        if pos is None or pos.qty < qty:
            have = Decimal("0") if pos is None else pos.qty
            raise ValueError(
                f"insufficient {side} qty for {code}/{currency}: need {qty}, "
                f"missing {qty - have}"
            )
        amortized_fee = pos.avg_fee_per_unit * qty
        avg_cost = pos.avg_cost
        mult = pos.multiplier
        first_open = pos.first_open_date
        note = pos.last_note
        pos.qty -= qty
        # qty 归零后清掉,避免后续看到陈旧的 avg
        if pos.qty == 0:
            del tab[key]
        return avg_cost, amortized_fee, mult, first_open, note

    def long_qty(self, code: str, currency: str) -> Decimal:
        pos = self._long.get((code, currency))
        return pos.qty if pos else Decimal("0")

    def short_qty(self, code: str, currency: str) -> Decimal:
        pos = self._short.get((code, currency))
        return pos.qty if pos else Decimal("0")

    def iter_open(self):
        """Yield (side, AvgPosition) for every still-open position. End-of-replay diagnostics."""
        for pos in self._long.values():
            yield "long", pos
        for pos in self._short.values():
            yield "short", pos

    # --- v1 接口兼容 (供既有 _handle_*_fallback 走 PositionLot 调用路径用) ---
    def open_lot(self, lot: PositionLot, side: str) -> None:
        self.open(
            code=lot.code, currency=lot.currency, side=side,
            qty=lot.quantity, price=lot.open_price,
            fee=lot.fee_per_unit * lot.quantity,
            open_date=lot.open_date, multiplier=lot.multiplier,
            note=lot.note,
        )

    def close_lot(self, code: str, currency: str, side: str,
                  quantity: Decimal):
        """兼容接口: 返回 v1 兼容的 [(伪 lot, used_qty, used_fee)] 单元素列表。"""
        avg_cost, amortized_fee, mult, first_open, note = self.close(code, currency, side, quantity)
        pseudo_lot = PositionLot(
            code=code, currency=currency,
            open_date=first_open, open_price=avg_cost, multiplier=mult,
            quantity=quantity, fee_per_unit=(amortized_fee / quantity if quantity else Decimal("0")),
            note=note,
        )
        return [(pseudo_lot, quantity, amortized_fee)]


# PositionBook:          Task 3 (done)
# --------------------------------------------------------------------------- #
# Trade event processor
# --------------------------------------------------------------------------- #


def _asset_type(trade: Trade) -> str:
    if trade.category == "期权":
        return "期权"
    if trade.category == "基金":
        return "基金"
    return "股票"


def _stock_underlying_for_option(code: str, currency: str) -> str:
    """从期权代码推导底层股票代码;失败回退 underlying 字母部分。"""
    try:
        meta = parse_option_code(code)
    except ValueError:
        return code
    if currency == "HKD" and meta.underlying in OPTION_UNDERLYING_TO_STOCK:
        return OPTION_UNDERLYING_TO_STOCK[meta.underlying]
    return meta.underlying


def _classify_stock_close(trade_note: str, trade_direction: str) -> str:
    """v2 股票卖出/平仓方式精细分类(基于 trade.note 中由期权事件注入的标识)。
    MWA 下建仓来源被合并, 只能按"卖出方式"区分:
      - CALL被指派卖出  : trade.note 含 'CALL 被指派'
      - PUT行权卖出     : trade.note 含 'PUT 行权'
      - 强制平仓        : trade.direction == '强平'
      - 普通市场卖出    : 其余主动 / 被动卖出
    PUT 被指派 与 CALL 行权 是"建仓"事件, 不进入此函数(走 open 分支)。
    """
    if trade_direction == "强平":
        return "强制平仓"
    n = trade_note or ""
    if "CALL 被指派" in n:
        return "CALL被指派卖出"
    if "PUT 行权" in n:
        return "PUT行权卖出"
    return "普通市场卖出"


def _classify_option_close(code: str, sub: str) -> str:
    """期权侧细分: 前缀加 PUT/CALL。sub ∈ {'到期作废', '主动平仓', '强制平仓'}。"""
    try:
        meta = parse_option_code(code)
        return f"{meta.opt_type and ('PUT' if meta.opt_type == 'P' else 'CALL')}{sub}"
    except ValueError:
        return sub


def _classify_trade_action(direction: str, quantity: Decimal) -> Optional[tuple[str, str]]:
    """Returns ('open'|'close', 'long'|'short') or None if the trade should be skipped."""
    if direction in ("买入开仓", "申购"):
        return ("open", "long")
    if direction == "卖出开仓":
        return ("open", "short")
    if direction in ("卖出平仓", "赎回"):
        return ("close", "long")
    if direction == "买入平仓":
        return ("close", "short")
    if direction == "强平":
        return ("close", "long") if quantity < 0 else ("close", "short")
    return None


def process_trade(trade: Trade, book: PositionBook) -> list[RealizedPnL]:
    """Process a single normalized Trade row, mutating book and returning realized P/L."""
    if trade.category in SKIP_CATEGORIES or trade.direction in SKIP_DIRECTIONS:
        return []
    action = _classify_trade_action(trade.direction, trade.quantity)
    if action is None:
        return []

    abs_qty = abs(trade.quantity)
    if abs_qty == 0:
        return []

    if action[0] == "open":
        side = action[1]
        fee_per_unit = trade.fee_total / abs_qty
        lot = PositionLot(
            code=trade.code, currency=trade.currency,
            open_date=trade.ts.date(),
            open_price=trade.price, multiplier=trade.multiplier,
            quantity=abs_qty, fee_per_unit=fee_per_unit,
            note=trade.note,
        )
        book.open_lot(lot, side=side)
        return []

    side = action[1]
    fee_per_unit_close = trade.fee_total / abs_qty
    pairs = book.close_lot(trade.code, trade.currency, side=side, quantity=abs_qty)
    realized: list[RealizedPnL] = []
    asset = _asset_type(trade)
    underlying = (
        _stock_underlying_for_option(trade.code, trade.currency)
        if asset == "期权" else trade.code
    )
    for lot, used_qty, used_open_fee in pairs:
        used_close_fee = fee_per_unit_close * used_qty
        if side == "long":
            gross = (trade.price - lot.open_price) * used_qty * trade.multiplier
            direction_label = "多头"
        else:
            gross = (lot.open_price - trade.price) * used_qty * trade.multiplier
            direction_label = "空头"
        pnl = gross - used_open_fee - used_close_fee
        if asset == "期权":
            base = "强制平仓" if trade.direction == "强平" else "主动平仓"
            sc = _classify_option_close(trade.code, base)
        else:
            sc = _classify_stock_close(trade.note, trade.direction)
        realized.append(RealizedPnL(
            realized_year=trade.ts.year,
            code=trade.code, asset_type=asset,
            direction=direction_label, quantity=used_qty,
            open_date=lot.open_date, open_price=lot.open_price,
            close_date=trade.ts.date(), close_price=trade.price,
            open_fee=used_open_fee, close_fee=used_close_fee,
            multiplier=trade.multiplier, currency=trade.currency,
            pnl=pnl, note=lot.note,
            underlying=underlying, sub_category=sc,
        ))
    return realized


# Trade event processor: Task 4 (done)
# --------------------------------------------------------------------------- #
# Option event handlers (asset-movement rows)
# --------------------------------------------------------------------------- #

_EVENT_RE = re.compile(r"^Opt (EXP-NA|ASS-P|ASS-C|ASS-PC|EXR)-")


def parse_asset_movement_event(note: str) -> str:
    """识别期权事件备注。富途账单同时存在中英两套写法 + ASS-PC 变体。

    返回标准化的事件代码:
      - EXP-NA  : "Opt EXP-NA-..." 或 "Option Expiration"
      - ASS-P   : "Opt ASS-P-..." / "Opt ASS-PC-..." 或 "Option Assignment"
                 (PUT/CALL 由合约代码自身的 P|C 分流)
      - EXR     : "Opt EXR-..."
    其余备注(Account Upgrade / Option Adjustment / DTC IN / Gift / 空 等)
    都不是期权事件,抛 ValueError 让调用方按需跳过。
    """
    s = note.strip()
    m = _EVENT_RE.match(s)
    if m:
        ev = m.group(1)
        if ev in ("ASS-P", "ASS-C", "ASS-PC"):
            return "ASS-P"  # 内部统一,后续按合约 P/C 分流
        if ev == "EXR":
            return "EXR"
        return "EXP-NA"
    if s == "Option Expiration":
        return "EXP-NA"
    if s == "Option Assignment":
        return "ASS-P"
    raise ValueError(f"unrecognized asset-movement note: {note!r}")


def process_option_event(mv: AssetMovement, book: PositionBook,
                         tracked_events: list[OptionEvent],
                         pending_index: Optional[dict] = None) -> list[RealizedPnL]:
    """Process one row from 证券-资产进出. Mutates book and tracked_events.

    pending_index: dict[(stock_code, currency, direction)] -> list[Trade refs]
    若提供且能匹配上同期股票流水,则修改对应 Trade 的 price 注入权利金调整;
    匹配失败时 fallback 到旧 spec 行为(自己动股票队列)。
    """
    if mv.event == "EXP-NA":
        return _handle_exp_na(mv, book, tracked_events)
    if mv.event in ("ASS-P", "ASS-C"):
        meta = parse_option_code(mv.code)
        if meta.opt_type == "P":
            return _handle_ass_put(mv, book, tracked_events, pending_index)
        return _handle_ass_call(mv, book, tracked_events, pending_index)
    if mv.event == "EXR":
        return _handle_exr(mv, book, tracked_events, pending_index)
    raise ValueError(f"unknown event: {mv.event}")


def _handle_exp_na(mv: AssetMovement, book: PositionBook,
                   tracked_events: list[OptionEvent]) -> list[RealizedPnL]:
    """Option expired worthless: close the entire queue side at price 0, fee 0."""
    short_q = book.short_qty(mv.code, mv.currency)
    long_q = book.long_qty(mv.code, mv.currency)
    if short_q >= mv.quantity and short_q > 0:
        side = "short"
        direction_label = "空头"
    elif long_q >= mv.quantity and long_q > 0:
        side = "long"
        direction_label = "多头"
    else:
        tracked_events.append(OptionEvent(
            code=mv.code, event_type="到期作废(无对应建仓)",
            event_date=mv.ts.date(), contracts=mv.quantity,
            strike=None, premium_net=None, cost_paid=None,
            transferred_to_stock=False,
            note=f"EXP-NA but no matching open lot (long={long_q}, short={short_q})",
        ))
        return []

    pairs = book.close_lot(mv.code, mv.currency, side=side, quantity=mv.quantity)
    realized: list[RealizedPnL] = []
    underlying = _stock_underlying_for_option(mv.code, mv.currency)
    sub_cat = _classify_option_close(mv.code, "到期作废")
    for lot, used_qty, used_open_fee in pairs:
        if side == "long":
            gross = (Decimal("0") - lot.open_price) * used_qty * lot.multiplier
        else:
            gross = (lot.open_price - Decimal("0")) * used_qty * lot.multiplier
        pnl = gross - used_open_fee
        realized.append(RealizedPnL(
            realized_year=mv.ts.year,
            code=mv.code, asset_type="期权", direction=direction_label,
            quantity=used_qty, open_date=lot.open_date, open_price=lot.open_price,
            close_date=mv.ts.date(), close_price=Decimal("0"),
            open_fee=used_open_fee, close_fee=Decimal("0"),
            multiplier=lot.multiplier, currency=mv.currency,
            pnl=pnl, note="期权到期作废",
            underlying=underlying, sub_category=sub_cat,
        ))
    tracked_events.append(OptionEvent(
        code=mv.code, event_type="到期作废", event_date=mv.ts.date(),
        contracts=mv.quantity, strike=None, premium_net=None, cost_paid=None,
        transferred_to_stock=False,
    ))
    return realized


from datetime import timedelta as _td  # for time window matching


def _option_underlying_stock_code(meta: OptionMeta, currency: str) -> str:
    """USD options use the ticker as-is; HKD options need mapping lookup."""
    if currency == "USD":
        return meta.underlying
    return meta.underlying_stock_code


def _try_match_and_adjust_trade(
    pending_index: Optional[dict],
    stock_code: str, currency: str, direction: str,
    target_qty_abs: Decimal, target_price: Decimal,
    mv_ts: datetime, adjusted_price: Decimal, note: str,
    window_days: int = 7,
) -> Optional["Trade"]:
    """在 pending_index 中查找匹配的未消费 trade,匹配则修改 price + note 并返回 trade。
    匹配条件: 同 (stock_code, currency, direction), |qty|=target_qty_abs,
    price≈target_price, ts 在 mv_ts ± window_days 内, 且未被任何 movement 标记消费。
    """
    if pending_index is None:
        return None
    key = (stock_code, currency, direction)
    candidates = pending_index.get(key, [])
    lo = mv_ts - _td(days=window_days)
    hi = mv_ts + _td(days=window_days)
    for t in candidates:
        if getattr(t, "_consumed_by_mv", False):
            continue
        if abs(t.quantity) != target_qty_abs:
            continue
        if abs(t.price - target_price) > Decimal("0.01"):
            continue
        if not (lo <= t.ts <= hi):
            continue
        t._consumed_by_mv = True  # type: ignore[attr-defined]
        t.price = adjusted_price
        t.note = note
        return t
    return None


def _sum_open_for_short(book: PositionBook, code: str, currency: str,
                        qty: Decimal) -> tuple[Decimal, Decimal, list]:
    pairs = book.close_lot(code, currency, side="short", quantity=qty)
    sum_open_amount = Decimal("0")
    sum_open_fee = Decimal("0")
    for lot, used_qty, used_fee in pairs:
        sum_open_amount += lot.open_price * used_qty * lot.multiplier
        sum_open_fee += used_fee
    return sum_open_amount, sum_open_fee, pairs


def _sum_open_for_long(book: PositionBook, code: str, currency: str,
                       qty: Decimal) -> tuple[Decimal, Decimal, list]:
    pairs = book.close_lot(code, currency, side="long", quantity=qty)
    sum_open_amount = Decimal("0")
    sum_open_fee = Decimal("0")
    for lot, used_qty, used_fee in pairs:
        sum_open_amount += lot.open_price * used_qty * lot.multiplier
        sum_open_fee += used_fee
    return sum_open_amount, sum_open_fee, pairs


def _handle_ass_put(mv: AssetMovement, book: PositionBook,
                    tracked_events: list[OptionEvent],
                    pending_index: Optional[dict] = None) -> list[RealizedPnL]:
    """卖出 PUT 被指派: 出空头 PUT 队列。优先调整同期"买入开仓"流水的价格;
    若找不到匹配流水,fallback 直接向股票多头队列入队。
    """
    meta = parse_option_code(mv.code)
    sum_open, sum_fee, pairs = _sum_open_for_short(book, mv.code, mv.currency, mv.quantity)
    premium_net = sum_open - sum_fee
    mult = pairs[0][0].multiplier if pairs else DEFAULT_OPTION_MULTIPLIER[mv.currency]
    shares = mv.quantity * mult
    stock_cost = meta.strike - premium_net / shares
    stock_code = _option_underlying_stock_code(meta, mv.currency)
    matched = _try_match_and_adjust_trade(
        pending_index, stock_code, mv.currency,
        direction="买入开仓", target_qty_abs=shares,
        target_price=meta.strike, mv_ts=mv.ts,
        adjusted_price=stock_cost,
        note=f"由 {mv.code} PUT 被指派调整建仓价(原行权价 {meta.strike} → {stock_cost})",
    )
    if matched is None:
        book.open_lot(PositionLot(
            code=stock_code, currency=mv.currency,
            open_date=mv.ts.date(),
            open_price=stock_cost, multiplier=Decimal("1"),
            quantity=shares, fee_per_unit=Decimal("0"),
            note=f"来自 {mv.code} PUT 被指派(无对应流水,fallback 入队)",
        ), side="long")
    tracked_events.append(OptionEvent(
        code=mv.code, event_type="PUT被指派", event_date=mv.ts.date(),
        contracts=mv.quantity, strike=meta.strike,
        premium_net=premium_net, cost_paid=None,
        transferred_to_stock=True,
        note=(f"调整流水: {shares} 股 {stock_code} @ {stock_cost}" if matched
              else f"无匹配流水, fallback 入队 {shares} 股 {stock_code} @ {stock_cost}"),
    ))
    return []


def _handle_ass_call(mv: AssetMovement, book: PositionBook,
                     tracked_events: list[OptionEvent],
                     pending_index: Optional[dict] = None) -> list[RealizedPnL]:
    """卖出 CALL 被指派: 出空头 CALL 队列。优先调整同期"卖出平仓"流水的价格;
    若找不到匹配流水,fallback FIFO 关闭股票多头队列(立即生成 RealizedPnL)。
    """
    meta = parse_option_code(mv.code)
    sum_open, sum_fee, opt_pairs = _sum_open_for_short(book, mv.code, mv.currency, mv.quantity)
    premium_net = sum_open - sum_fee
    mult = opt_pairs[0][0].multiplier if opt_pairs else DEFAULT_OPTION_MULTIPLIER[mv.currency]
    shares = mv.quantity * mult
    sell_price = meta.strike + premium_net / shares
    stock_code = _option_underlying_stock_code(meta, mv.currency)
    matched = _try_match_and_adjust_trade(
        pending_index, stock_code, mv.currency,
        direction="卖出平仓", target_qty_abs=shares,
        target_price=meta.strike, mv_ts=mv.ts,
        adjusted_price=sell_price,
        note=f"由 {mv.code} CALL 被指派调整平仓价(原行权价 {meta.strike} → {sell_price})",
    )
    realized: list[RealizedPnL] = []
    if matched is None:
        # fallback: spec §5.3 旧逻辑,直接 FIFO 关股票多头
        stock_pairs = book.close_lot(stock_code, mv.currency, side="long", quantity=shares)
        for lot, used_qty, used_open_fee in stock_pairs:
            gross = (sell_price - lot.open_price) * used_qty * Decimal("1")
            pnl = gross - used_open_fee
            realized.append(RealizedPnL(
                realized_year=mv.ts.year,
                code=stock_code, asset_type="股票", direction="多头",
                quantity=used_qty, open_date=lot.open_date, open_price=lot.open_price,
                close_date=mv.ts.date(), close_price=sell_price,
                open_fee=used_open_fee, close_fee=Decimal("0"),
                multiplier=Decimal("1"), currency=mv.currency, pnl=pnl,
                note=f"CALL被指派 (来自 {mv.code}, fallback)",
                underlying=stock_code, sub_category="CALL被指派卖出(fallback)",
            ))
    tracked_events.append(OptionEvent(
        code=mv.code, event_type="CALL被指派", event_date=mv.ts.date(),
        contracts=mv.quantity, strike=meta.strike,
        premium_net=premium_net, cost_paid=None,
        transferred_to_stock=True,
        note=(f"调整流水: 卖出 {shares} 股 {stock_code} @ {sell_price}" if matched
              else f"无匹配流水, fallback 关 {shares} 股 {stock_code} @ {sell_price}"),
    ))
    return realized


def _handle_exr(mv: AssetMovement, book: PositionBook,
                tracked_events: list[OptionEvent],
                pending_index: Optional[dict] = None) -> list[RealizedPnL]:
    """自行行权: PUT → 调整"卖出平仓"流水(或 fallback 关股票多头);
    CALL → 调整"买入开仓"流水(或 fallback 入股票多头)。
    """
    meta = parse_option_code(mv.code)
    sum_open, sum_fee, opt_pairs = _sum_open_for_long(book, mv.code, mv.currency, mv.quantity)
    cost_paid = sum_open + sum_fee
    mult = opt_pairs[0][0].multiplier if opt_pairs else DEFAULT_OPTION_MULTIPLIER[mv.currency]
    shares = mv.quantity * mult
    stock_code = _option_underlying_stock_code(meta, mv.currency)
    realized: list[RealizedPnL] = []
    if meta.opt_type == "P":
        sell_price = meta.strike - cost_paid / shares
        matched = _try_match_and_adjust_trade(
            pending_index, stock_code, mv.currency,
            direction="卖出平仓", target_qty_abs=shares,
            target_price=meta.strike, mv_ts=mv.ts,
            adjusted_price=sell_price,
            note=f"由 {mv.code} PUT 行权调整平仓价(原行权价 {meta.strike} → {sell_price})",
        )
        if matched is None:
            stock_pairs = book.close_lot(stock_code, mv.currency, side="long", quantity=shares)
            for lot, used_qty, used_open_fee in stock_pairs:
                gross = (sell_price - lot.open_price) * used_qty
                pnl = gross - used_open_fee
                realized.append(RealizedPnL(
                    realized_year=mv.ts.year,
                    code=stock_code, asset_type="股票", direction="多头",
                    quantity=used_qty, open_date=lot.open_date, open_price=lot.open_price,
                    close_date=mv.ts.date(), close_price=sell_price,
                    open_fee=used_open_fee, close_fee=Decimal("0"),
                    multiplier=Decimal("1"), currency=mv.currency, pnl=pnl,
                    note=f"PUT行权 (来自 {mv.code}, fallback)",
                    underlying=stock_code, sub_category="PUT行权卖出(fallback)",
                ))
        tracked_events.append(OptionEvent(
            code=mv.code, event_type="PUT行权", event_date=mv.ts.date(),
            contracts=mv.quantity, strike=meta.strike,
            premium_net=None, cost_paid=cost_paid,
            transferred_to_stock=True,
            note=(f"调整流水: 卖出 {shares} 股 {stock_code} @ {sell_price}" if matched
                  else f"无匹配流水, fallback 关 {shares} 股 {stock_code} @ {sell_price}"),
        ))
        return realized
    else:  # CALL
        buy_price = meta.strike + cost_paid / shares
        matched = _try_match_and_adjust_trade(
            pending_index, stock_code, mv.currency,
            direction="买入开仓", target_qty_abs=shares,
            target_price=meta.strike, mv_ts=mv.ts,
            adjusted_price=buy_price,
            note=f"由 {mv.code} CALL 行权调整建仓价(原行权价 {meta.strike} → {buy_price})",
        )
        if matched is None:
            book.open_lot(PositionLot(
                code=stock_code, currency=mv.currency,
                open_date=mv.ts.date(),
                open_price=buy_price, multiplier=Decimal("1"),
                quantity=shares, fee_per_unit=Decimal("0"),
                note=f"来自 {mv.code} CALL 行权(无对应流水, fallback)",
            ), side="long")
        tracked_events.append(OptionEvent(
            code=mv.code, event_type="CALL行权", event_date=mv.ts.date(),
            contracts=mv.quantity, strike=meta.strike,
            premium_net=None, cost_paid=cost_paid,
            transferred_to_stock=True,
            note=(f"调整流水: 买入 {shares} 股 {stock_code} @ {buy_price}" if matched
                  else f"无匹配流水, fallback 入队 {shares} 股 {stock_code} @ {buy_price}"),
        ))
        return realized


# Option event handlers: Task 5/6 (done)
# --------------------------------------------------------------------------- #
# Excel loader
# --------------------------------------------------------------------------- #

import openpyxl  # noqa: E402

YEAR_FILE_RE = re.compile(r"^(\d{4})_年度账单_.*\.xlsx$")


def _to_decimal(v) -> Decimal:
    if v is None or v == "" or v == "-":
        return Decimal("0")
    return Decimal(str(v))


def _to_datetime(v) -> datetime:
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    s = str(v).strip()
    if re.fullmatch(r"\d{8}", s):
        return datetime(int(s[:4]), int(s[4:6]), int(s[6:8]))
    return datetime.fromisoformat(s)


def load_year(xlsx_path: Path) -> dict:
    """Load one year file. Returns dict with keys: trades, movements, multipliers, seed_lots."""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    multipliers: dict[tuple[str, str], Decimal] = {}
    seed_lots: list[tuple[date, PositionLot, str]] = []
    position_snapshots: list[dict] = []
    ws = wb["证券-持仓总览"]
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = row
            continue
        rec = dict(zip(header, row))
        if str(rec.get("账户号码") or "") not in TARGET_ACCOUNTS:
            continue
        code = rec.get("代码名称")
        currency = rec.get("币种")
        mult = _to_decimal(rec.get("乘数"))
        if code and currency and mult > 0:
            multipliers[(str(code), str(currency))] = mult
        if rec.get("时期类型") == "期初":
            qty = _to_decimal(rec.get("数量/面值"))
            if qty == 0:
                continue
            snap_date = _to_datetime(rec.get("日期")).date()
            price = _to_decimal(rec.get("价格"))
            side = "long" if qty > 0 else "short"
            lot = PositionLot(
                code=str(code), currency=str(currency),
                open_date=snap_date, open_price=price,
                multiplier=mult if mult > 0 else Decimal("1"),
                quantity=abs(qty), fee_per_unit=Decimal("0"),
                note=f"种子({snap_date.year}期初)",
            )
            seed_lots.append((snap_date, lot, side))
        elif rec.get("时期类型") == "期末":
            snap_date = _to_datetime(rec.get("日期")).date()
            qty = _to_decimal(rec.get("数量/面值"))
            price = _to_decimal(rec.get("价格"))
            mv = _to_decimal(rec.get("市值"))
            side = "long" if qty > 0 else "short"
            position_snapshots.append({
                "date": snap_date,
                "code": str(code), "currency": str(currency),
                "qty": abs(qty), "price": price,
                "market_value": mv, "side": side,
            })

    trades: list[Trade] = []
    ws = wb["证券-交易流水"]
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = row
            continue
        rec = dict(zip(header, row))
        if str(rec.get("账户号码") or "") not in TARGET_ACCOUNTS:
            continue
        category = str(rec.get("品类") or "")
        if category in SKIP_CATEGORIES:
            continue
        direction = str(rec.get("方向") or "")
        if direction in SKIP_DIRECTIONS:
            continue
        currency = str(rec.get("币种") or "")
        code = str(rec.get("代码名称") or "")
        mult_key = (code, currency)
        if mult_key in multipliers:
            mult = multipliers[mult_key]
        elif category == "期权":
            mult = DEFAULT_OPTION_MULTIPLIER.get(currency, Decimal("100"))
        else:
            mult = Decimal("1")
        trades.append(Trade(
            ts=_to_datetime(rec.get("成交时间")),
            category=category, code=code, market=str(rec.get("交易所/市场") or ""),
            currency=currency, direction=direction,
            quantity=_to_decimal(rec.get("数量/面值")),
            price=_to_decimal(rec.get("价格")),
            fee_total=_to_decimal(rec.get("总费用")),
            amount_change=_to_decimal(rec.get("变动金额")),
            multiplier=mult,
        ))

    movements: list[AssetMovement] = []
    adjustments: list[tuple[date, str, Decimal, str]] = []  # (日期, 代码, 带符号数量, 备注)
    other_unhandled: list[tuple[date, str, str, Decimal, str]] = []  # warning 用
    ws = wb["证券-资产进出"]
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = row
            continue
        rec = dict(zip(header, row))
        if str(rec.get("账户号码") or "") not in TARGET_ACCOUNTS:
            continue
        category = str(rec.get("品类") or "")
        if category == "基金":
            continue
        note = str(rec.get("备注") or "")
        d = _to_datetime(rec.get("日期")).date()
        code = str(rec.get("代码名称") or "")
        qty_signed = _to_decimal(rec.get("数量"))
        currency = str(rec.get("币种") or "")
        # Option Adjustment(代码改名/拆股),单独收集供 alias 推断
        if "Option Adjustment" in note and category == "期权":
            adjustments.append((d, code, qty_signed, note))
            continue
        # 多空反转平仓(空备注 + 类型=多空反转平仓): In/Out 同代码同量, 净影响 0, 跳过
        if str(rec.get("类型") or "") == "多空反转平仓":
            continue
        if category != "期权":
            # 股票相关的资产进出(DTC IN / Gift / Account Upgrade) 暂不参与回放,仅记 warning
            if note and note != "Account Upgrade":
                other_unhandled.append((d, code, currency, qty_signed, note))
            continue
        try:
            event = parse_asset_movement_event(note)
        except ValueError:
            if note and note != "Account Upgrade":
                other_unhandled.append((d, code, currency, qty_signed, note))
            continue
        movements.append(AssetMovement(
            ts=_to_datetime(rec.get("日期")),
            code=code,
            currency=currency,
            quantity=abs(qty_signed),  # 方向由 event 类型决定,数量取绝对值
            note=note, event=event,
        ))

    return {"trades": trades, "movements": movements,
            "multipliers": multipliers, "seed_lots": seed_lots,
            "adjustments": adjustments, "other_unhandled": other_unhandled,
            "position_snapshots": position_snapshots}


def _build_alias_map(adjustments: list[tuple[date, str, Decimal, str]]
                     ) -> tuple[dict[str, str], list[str]]:
    """从 Option Adjustment 配对推断代码别名映射(out_code -> in_code)。

    1:1 配对(同日,绝对值相等的 OUT + IN)视为重命名;
    数量不匹配的(NVDA 拆股 1:10 等)记入 unresolved 警告,不进 alias_map。
    """
    alias: dict[str, str] = {}
    unresolved: list[str] = []
    by_day: dict[date, list[tuple[str, Decimal, str]]] = defaultdict(list)
    for d, code, qty, note in adjustments:
        by_day[d].append((code, qty, note))
    for d, items in by_day.items():
        outs = [(c, abs(q), note) for c, q, note in items if q < 0]
        ins = [(c, abs(q), note) for c, q, note in items if q > 0]
        used_in: set[int] = set()
        for oc, oq, onote in outs:
            paired = False
            for idx, (ic, iq, inote) in enumerate(ins):
                if idx in used_in:
                    continue
                if iq == oq and onote == inote:
                    alias[oc] = ic
                    used_in.add(idx)
                    paired = True
                    break
            if not paired:
                unresolved.append(f"{d} OUT {oc} qty={oq} note={onote!r} 无 1:1 IN 配对")
        for idx, (ic, iq, inote) in enumerate(ins):
            if idx not in used_in:
                unresolved.append(f"{d} IN {ic} qty={iq} note={inote!r} 无 1:1 OUT 配对")
    return alias, unresolved


def _apply_alias_to_code(code: str, alias: dict[str, str]) -> str:
    """递归应用别名(防止 A->B->C 链)。"""
    seen = set()
    cur = code
    while cur in alias and cur not in seen:
        seen.add(cur)
        cur = alias[cur]
    return cur


def _parse_overrides_from_dict(raw: dict) -> list[dict]:
    """从配置字典中解析 overrides 列表。

    overrides 每条形如:
      {"code": "BABA", "currency": "USD",
       "dtc_in": [{"date": "YYYY-MM-DD", "qty": "<str>", "price": "<str>", "note": "..."}]}
    """
    out: list[dict] = []
    for item in raw.get("overrides", []):
        dtc = []
        for r in item.get("dtc_in", []):
            dtc.append({
                "date": datetime.strptime(r["date"], "%Y-%m-%d").date(),
                "qty": Decimal(str(r["qty"])),
                "price": Decimal(str(r["price"])),
                "note": r.get("note", ""),
            })
        dtc.sort(key=lambda r: r["date"])
        out.append({
            "code": item["code"], "currency": item["currency"], "dtc_in": dtc,
        })
    return out


def _parse_tax_config_from_dict(raw: dict) -> dict:
    """从配置字典中解析税务配置。

    可选字段:
      "tax_rate": "0.20"           — 覆盖默认税率
      "fx_rates": {"2023": "7.0827", "2024": "7.1884"}  — 覆盖年末 USD/CNY 汇率
    """
    tax_config: dict = {}
    if "tax_rate" in raw:
        tax_config["tax_rate"] = Decimal(str(raw["tax_rate"]))
    if "fx_rates" in raw:
        tax_config["fx_rates"] = {
            int(k): Decimal(str(v)) for k, v in raw["fx_rates"].items()
        }
    return tax_config


def load_legacy_overrides(path: Path) -> dict:
    """读取旧版 cost_basis_overrides.json (向后兼容)。

    返回含 overrides / tax_config 的配置字典。
    如果文件不存在,返回空字典。
    """
    empty = {"overrides": [], "tax_config": {}}
    if not path.exists():
        return empty
    raw = json.loads(path.read_text(encoding="utf-8"))
    overrides = _parse_overrides_from_dict(raw)
    tax_config = _parse_tax_config_from_dict(raw)
    return {"overrides": overrides, "tax_config": tax_config}


def apply_cost_basis_overrides(
    overrides: list[dict],
    per_year: list[tuple[int, Path, dict]],
    replay_year: int,
    seed_lots: list[tuple[date, PositionLot, str]],
    all_trades: list[Trade],
    multipliers: dict[tuple[str, str], Decimal],
) -> list[str]:
    """根据用户提供的 DTC IN 成本基础,重算种子持仓 avg_cost 并注入 replay 期内 DTC IN trade。

    返回 info 列表(供 warnings 区展示),记录每条 override 的处理结果。
    """
    info: list[str] = []
    # 收集 < replay_year 的全部 trades(原始,未参与回放),用于在重算 pre-replay avg 时扣除卖出
    pre_trades_by_key: dict[tuple[str, str], list[Trade]] = defaultdict(list)
    for y, _, data in per_year:
        if y >= replay_year:
            continue
        for t in data["trades"]:
            if t.category != "证券":
                continue
            pre_trades_by_key[(t.code, t.currency)].append(t)

    for ov in overrides:
        code, ccy = ov["code"], ov["currency"]
        dtc = ov["dtc_in"]
        if not dtc:
            continue
        pre_dtc = [r for r in dtc if r["date"].year < replay_year]
        post_dtc = [r for r in dtc if r["date"].year >= replay_year]

        # ---- 1. 重算种子持仓 ----
        if pre_dtc:
            # 把 pre-replay DTC IN 与该期间证券 trades 按时间排序后逐条 MWA
            events: list[tuple[datetime, str, Decimal, Decimal]] = []
            for r in pre_dtc:
                events.append((datetime.combine(r["date"], datetime.min.time()),
                               "in", r["qty"], r["price"]))
            for t in pre_trades_by_key.get((code, ccy), []):
                if t.direction in ("买入开仓", "卖出平仓", "强平"):
                    events.append((t.ts, "trade", t.quantity, t.price))
            events.sort(key=lambda e: e[0])
            qty, avg = Decimal("0"), Decimal("0")
            for _, kind, q, p in events:
                if kind == "in":
                    new_qty = qty + q
                    avg = (avg * qty + p * q) / new_qty if new_qty > 0 else Decimal("0")
                    qty = new_qty
                else:
                    # MWA: 买入加权累加, 卖出不动 avg
                    if q > 0:
                        new_qty = qty + q
                        avg = (avg * qty + p * q) / new_qty if new_qty > 0 else Decimal("0")
                        qty = new_qty
                    else:
                        qty += q  # q<0
                        if qty <= 0:
                            qty, avg = Decimal("0"), Decimal("0")
            # 替换种子 lot
            replaced = False
            for idx, (sd, lot, side) in enumerate(seed_lots):
                if lot.code == code and lot.currency == ccy and side == "long":
                    if lot.quantity != qty:
                        info.append(
                            f"[override] {code}/{ccy} 种子数量不匹配: "
                            f"快照 {lot.quantity} vs override 推算 {qty}, 仍按快照数量{lot.quantity}保留,"
                            f"open_price 替换为 {avg:.6f}"
                        )
                    else:
                        info.append(
                            f"[override] {code}/{ccy} 种子 open_price 替换: "
                            f"{lot.open_price} → {avg:.6f} (按 {len(pre_dtc)} 笔 DTC IN MWA)"
                        )
                    lot.open_price = avg
                    lot.note = (lot.note or "") + f" [override avg={avg:.6f}]"
                    replaced = True
                    break
            if not replaced:
                info.append(
                    f"[override] {code}/{ccy} 期初快照中无该多头持仓,"
                    f"pre-replay DTC IN 已忽略(若需要,可手工增加种子 lot)"
                )

        # ---- 2. 注入 >= replay_year 的 DTC IN 为买入开仓 trade ----
        mult = multipliers.get((code, ccy), Decimal("1"))
        for r in post_dtc:
            all_trades.append(Trade(
                ts=datetime.combine(r["date"], datetime.min.time()),
                category="证券", code=code, market="",
                currency=ccy, direction="买入开仓",
                quantity=r["qty"], price=r["price"],
                fee_total=Decimal("0"),
                amount_change=-r["qty"] * r["price"],
                multiplier=mult,
                note=f"DTC IN override: {r['note']}",
            ))
            info.append(
                f"[override] {code}/{ccy} 注入买入开仓: "
                f"{r['date']} qty={r['qty']} price={r['price']}"
            )
    return info


def load_all_years(target_dir: Path) -> dict:
    """Scan YYYY_年度账单_*.xlsx files in target_dir. Returns merged data.

    回放起点 = 第一个含"期初"快照的年份。早于该年的文件仅参与多头乘数登记,
    其交易与资产进出会被丢弃(否则状态不闭合)。

    完成跨年聚合后,把 Option Adjustment 推断出的代码别名(BABA1<X>->BABA<X> 等)
    应用到所有 trades / movements / multipliers / seed_lots 上,使 FIFO 配对一致。
    """
    files: list[tuple[int, Path]] = []
    for p in target_dir.glob("*_年度账单_*.xlsx"):
        m = YEAR_FILE_RE.match(p.name)
        if m:
            files.append((int(m.group(1)), p))
    if not files:
        raise FileNotFoundError(f"no YYYY_年度账单_*.xlsx files in {target_dir}")
    files.sort()
    per_year: list[tuple[int, Path, dict]] = []
    multipliers: dict[tuple[str, str], Decimal] = {}
    all_adjustments: list[tuple[date, str, Decimal, str]] = []
    all_other_unhandled: list = []
    all_snapshots: list[dict] = []
    for year, path in files:
        data = load_year(path)
        per_year.append((year, path, data))
        multipliers.update(data["multipliers"])
        all_adjustments.extend(data.get("adjustments", []))
        all_other_unhandled.extend(data.get("other_unhandled", []))
        for snap in data.get("position_snapshots", []):
            all_snapshots.append((year, snap))
    replay_year = next((y for y, _, d in per_year if d["seed_lots"]), None)
    if replay_year is None:
        # 无期初快照：用最早年份的期末快照作为种子，并跳过该年交易（避免重复计算）
        earliest_year, _, earliest_data = per_year[0]
        fallback_seeds: list[tuple[date, PositionLot, str]] = []
        for snap in earliest_data.get("position_snapshots", []):
            if snap["qty"] == 0:
                continue
            mult = multipliers.get((snap["code"], snap["currency"]), Decimal("1"))
            lot = PositionLot(
                code=snap["code"], currency=snap["currency"],
                open_date=snap["date"], open_price=snap["price"],
                multiplier=mult,
                quantity=snap["qty"], fee_per_unit=Decimal("0"),
                note=f"种子({snap['date'].year}期末-无期初快照)",
            )
            fallback_seeds.append((snap["date"], lot, snap["side"]))
        if len(per_year) == 1:
            # 只有一年账单：种子就是期末持仓，无交易回放
            replay_year = earliest_year
            seed_lots = fallback_seeds
            skipped_pre_years: list[int] = []
        else:
            # 多年账单：用最早年份期末作种子，从下一年开始回放
            replay_year = earliest_year + 1
            seed_lots = fallback_seeds
            skipped_pre_years = [earliest_year]
    else:
        seed_lots = next(d["seed_lots"] for y, _, d in per_year if y == replay_year)
        skipped_pre_years = [y for y, _, _ in per_year if y < replay_year]
    all_trades: list[Trade] = []
    all_movements: list[AssetMovement] = []
    for year, _, data in per_year:
        if year < replay_year:
            continue
        all_trades.extend(data["trades"])
        all_movements.extend(data["movements"])

    # 别名归一化
    alias, alias_unresolved = _build_alias_map(all_adjustments)
    for t in all_trades:
        t.code = _apply_alias_to_code(t.code, alias)
    for mv in all_movements:
        mv.code = _apply_alias_to_code(mv.code, alias)
    # multipliers / seed_lots 也对齐
    new_multipliers: dict[tuple[str, str], Decimal] = {}
    for (code, ccy), m in multipliers.items():
        new_multipliers[(_apply_alias_to_code(code, alias), ccy)] = m
    multipliers = new_multipliers
    new_seeds: list[tuple[date, PositionLot, str]] = []
    for sd, lot, side in seed_lots:
        lot.code = _apply_alias_to_code(lot.code, alias)
        new_seeds.append((sd, lot, side))
    seed_lots = new_seeds
    for _, snap in all_snapshots:
        snap["code"] = _apply_alias_to_code(snap["code"], alias)

    # 优先使用 script_config.json 中的 overrides,
    # 若存在旧版 cost_basis_overrides.json 则合并(后者优先级更高)
    overrides = list(_CONFIG_OVERRIDES)
    tax_config = dict(_CONFIG_TAX)
    legacy = load_legacy_overrides(target_dir / "cost_basis_overrides.json")
    if legacy["overrides"]:
        overrides = legacy["overrides"]
    if legacy["tax_config"]:
        tax_config.update(legacy["tax_config"])

    override_info = apply_cost_basis_overrides(
        overrides, per_year, replay_year, seed_lots, all_trades, multipliers,
    )

    return {
        "trades": all_trades, "movements": all_movements,
        "multipliers": multipliers, "seed_lots": seed_lots,
        "year_range": (replay_year, files[-1][0]),
        "skipped_pre_years": skipped_pre_years,
        "alias_map": alias,
        "alias_unresolved": alias_unresolved,
        "other_unhandled": all_other_unhandled,
        "files": [str(p) for _, p in files],
        "cost_basis_overrides": override_info,
        "position_snapshots": all_snapshots,
        "tax_config": tax_config,
    }


# --------------------------------------------------------------------------- #
# Replay loop
# --------------------------------------------------------------------------- #


def _snapshot_book(book: PositionBook) -> list[tuple[str, dict]]:
    """Capture current book state as a list of (side, position_dict)."""
    snap = []
    for side, pos in book.iter_open():
        snap.append((side, {
            "code": pos.code,
            "currency": pos.currency,
            "qty": pos.qty,
            "avg_cost": pos.avg_cost,
            "multiplier": pos.multiplier,
            "first_open_date": pos.first_open_date,
            "last_note": pos.last_note,
        }))
    return snap


def replay(trades: list[Trade], movements: list[AssetMovement],
           seed_lots: list[tuple[date, PositionLot, str]],
           ) -> tuple[list[RealizedPnL], list[OptionEvent], PositionBook, list[str], dict[int, list]]:
    """Walk events chronologically, returning realized P/L, tracked option events,
    end-state book, a list of warning strings, and year-end book snapshots.

    在期权事件触发时,优先调整同期股票流水的价格(注入权利金调整),
    使被动买卖与后续主动买卖在同一 FIFO 队列中对齐成本。
    遇到仓位不足/状态异常时记入 warnings 并跳过该事件,不中断回放。
    """
    book = PositionBook()
    for _, lot, side in seed_lots:
        book.open_lot(lot, side=side)

    pending_index: dict[tuple[str, str, str], list[Trade]] = defaultdict(list)
    for t in trades:
        if t.category == "证券":
            pending_index[(t.code, t.currency, t.direction)].append(t)

    events: list[tuple[datetime, int, object]] = []
    for t in trades:
        events.append((t.ts, 1, t))
    for mv in movements:
        events.append((mv.ts, 0, mv))
    events.sort(key=lambda e: (e[0], e[1]))

    realized: list[RealizedPnL] = []
    option_events: list[OptionEvent] = []
    warnings: list[str] = []
    year_end_books: dict[int, list] = {}
    current_year: int | None = None
    for _ts, _kind, ev in events:
        event_year = _ts.year if hasattr(_ts, 'year') else _ts.date().year if hasattr(_ts, 'date') else None
        if current_year is not None and event_year != current_year:
            year_end_books[current_year] = _snapshot_book(book)
        current_year = event_year
        try:
            if isinstance(ev, Trade):
                realized.extend(process_trade(ev, book))
            else:
                realized.extend(process_option_event(ev, book, option_events, pending_index))
        except (ValueError, KeyError) as e:
            ev_desc = (
                f"trade {ev.ts} {ev.code}/{ev.currency} {ev.direction} qty={ev.quantity} px={ev.price}"
                if isinstance(ev, Trade) else
                f"movement {ev.ts.date()} {ev.code}/{ev.currency} {ev.event} qty={ev.quantity}"
            )
            warnings.append(f"{ev_desc} -> 跳过: {e}")
    if current_year is not None:
        year_end_books[current_year] = _snapshot_book(book)
    return realized, option_events, book, warnings, year_end_books


# Excel loader / Replay: Task 7 (done)
# HTML / JSON / CLI 见文件末尾


# --------------------------------------------------------------------------- #
# Annual summary computation
# --------------------------------------------------------------------------- #


def _snaps_for_year(snapshots: list[tuple[int, dict]], year: int) -> list[dict]:
    return [s for y, s in snapshots if y == year]


def compute_annual_summary(
    trades: list[Trade],
    realized: list[RealizedPnL],
    position_snapshots: list[tuple[int, dict]],
    year_end_books: dict[int, list],
    year_range: tuple[int, int],
) -> list[dict]:
    """Compute annual summary: buy/sell volumes, start/end positions, realized & unrealized PnL."""

    # Year-start market value = previous year's 期末 market value (by currency)
    year_start_mv: dict[int, dict[str, Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal("0")))
    for snap_year, snap in position_snapshots:
        year_start_mv[snap_year + 1][snap["currency"]] += snap["market_value"]

    # Buy/sell totals per year/currency
    buy_totals: dict[tuple[int, str], Decimal] = defaultdict(lambda: Decimal("0"))
    sell_totals: dict[tuple[int, str], Decimal] = defaultdict(lambda: Decimal("0"))
    buy_counts: dict[tuple[int, str], int] = defaultdict(int)
    sell_counts: dict[tuple[int, str], int] = defaultdict(int)

    for t in trades:
        yr = t.ts.year
        amt = abs(t.price * t.quantity * t.multiplier)
        if t.direction in ("买入开仓", "申购"):
            buy_totals[(yr, t.currency)] += amt
            buy_counts[(yr, t.currency)] += 1
        elif t.direction in ("卖出平仓", "赎回"):
            sell_totals[(yr, t.currency)] += amt
            sell_counts[(yr, t.currency)] += 1
        elif t.direction == "卖出开仓":
            sell_totals[(yr, t.currency)] += amt
            sell_counts[(yr, t.currency)] += 1
        elif t.direction == "买入平仓":
            buy_totals[(yr, t.currency)] += amt
            buy_counts[(yr, t.currency)] += 1
        elif t.direction == "强平":
            if t.quantity < 0:
                # 强平多头 → 卖出
                sell_totals[(yr, t.currency)] += amt
                sell_counts[(yr, t.currency)] += 1
            else:
                # 强平空头 → 买入
                buy_totals[(yr, t.currency)] += amt
                buy_counts[(yr, t.currency)] += 1

    # Realized PnL by year/currency
    realized_by_year: dict[tuple[int, str], Decimal] = defaultdict(lambda: Decimal("0"))
    for r in realized:
        realized_by_year[(r.realized_year, r.currency)] += r.pnl

    # Collect all (year, currency) pairs
    all_yc: set[tuple[int, str]] = set()
    for (yr, ccy) in buy_totals:
        all_yc.add((yr, ccy))
    for (yr, ccy) in sell_totals:
        all_yc.add((yr, ccy))
    for yr in year_end_books:
        for _, pos in year_end_books[yr]:
            all_yc.add((yr, pos["currency"]))
    for yr in year_start_mv:
        for ccy in year_start_mv[yr]:
            all_yc.add((yr, ccy))

    # Filter to replay years
    all_yc = {(y, c) for y, c in all_yc if year_range[0] <= y <= year_range[1]}

    summary_rows = []
    for yr, ccy in sorted(all_yc):
        # Year-end positions from book
        end_positions = []
        if yr in year_end_books:
            for side, pos in year_end_books[yr]:
                if pos["currency"] == ccy and pos["qty"] != 0:
                    end_positions.append({
                        "code": pos["code"],
                        "qty": str(pos["qty"]),
                        "avg_cost": str(pos["avg_cost"]),
                        "side": side,
                    })

        # Year-end market values from snapshot
        snaps = _snaps_for_year(position_snapshots, yr)
        snap_by_code: dict[str, dict] = {}
        for s in snaps:
            if s["currency"] == ccy:
                snap_by_code[s["code"]] = s

        end_mv = sum((s["market_value"] for s in snaps if s["currency"] == ccy), Decimal("0"))

        # Unrealized PnL: (market_price - avg_cost) * qty for longs, reversed for shorts
        unrealized = Decimal("0")
        if yr in year_end_books:
            for side, pos in year_end_books[yr]:
                if pos["currency"] == ccy and pos["code"] in snap_by_code:
                    snap = snap_by_code[pos["code"]]
                    if side == "long":
                        unrealized += (snap["price"] - pos["avg_cost"]) * pos["qty"] * pos["multiplier"]
                    else:
                        unrealized += (pos["avg_cost"] - snap["price"]) * pos["qty"] * pos["multiplier"]

        # Year-start positions from previous year's 期末 snapshot
        start_snaps = _snaps_for_year(position_snapshots, yr - 1)
        start_count = sum(1 for s in start_snaps if s["currency"] == ccy)
        start_mv = year_start_mv.get(yr, {}).get(ccy, Decimal("0"))

        summary_rows.append({
            "year": yr,
            "currency": ccy,
            "start_count": start_count,
            "start_market_value": str(start_mv),
            "buy_count": buy_counts.get((yr, ccy), 0),
            "buy_amount": str(buy_totals.get((yr, ccy), Decimal("0"))),
            "sell_count": sell_counts.get((yr, ccy), 0),
            "sell_amount": str(sell_totals.get((yr, ccy), Decimal("0"))),
            "realized_pnl": str(realized_by_year.get((yr, ccy), Decimal("0"))),
            "end_count": len(end_positions),
            "end_market_value": str(end_mv),
            "unrealized_pnl": str(unrealized),
            "end_positions": end_positions,
        })

    # 添加每年的港币合计行（USD × 7.8 换算）
    HKD_PEG = Decimal("7.8")
    all_years = sorted({r["year"] for r in summary_rows})
    for yr in all_years:
        rows_this_year = [r for r in summary_rows if r["year"] == yr]
        if len(rows_this_year) <= 1:
            continue
        total_start_mv = Decimal("0")
        total_buy_amt = Decimal("0")
        total_sell_amt = Decimal("0")
        total_rpnl = Decimal("0")
        total_end_mv = Decimal("0")
        total_upnl = Decimal("0")
        total_buy_cnt = 0
        total_sell_cnt = 0
        total_start_cnt = 0
        total_end_cnt = 0
        for r in rows_this_year:
            factor = HKD_PEG if r["currency"] == "USD" else Decimal("1")
            total_start_cnt += r["start_count"]
            total_start_mv += Decimal(r["start_market_value"]) * factor
            total_buy_cnt += r["buy_count"]
            total_buy_amt += Decimal(r["buy_amount"]) * factor
            total_sell_cnt += r["sell_count"]
            total_sell_amt += Decimal(r["sell_amount"]) * factor
            total_rpnl += Decimal(r["realized_pnl"]) * factor
            total_end_cnt += r["end_count"]
            total_end_mv += Decimal(r["end_market_value"]) * factor
            total_upnl += Decimal(r["unrealized_pnl"]) * factor
        summary_rows.append({
            "year": yr,
            "currency": "HKD合计",
            "start_count": total_start_cnt,
            "start_market_value": str(total_start_mv),
            "buy_count": total_buy_cnt,
            "buy_amount": str(total_buy_amt),
            "sell_count": total_sell_cnt,
            "sell_amount": str(total_sell_amt),
            "realized_pnl": str(total_rpnl),
            "end_count": total_end_cnt,
            "end_market_value": str(total_end_mv),
            "unrealized_pnl": str(total_upnl),
            "end_positions": [],
        })

    return summary_rows


# --------------------------------------------------------------------------- #
# HTML render
# --------------------------------------------------------------------------- #


def _fmt_dec(d: Decimal, places: int = 2) -> str:
    q = Decimal(10) ** -places
    return str(d.quantize(q, rounding=ROUND_HALF_UP))


def _esc(s) -> str:
    return html.escape(str(s)) if s is not None else ""


def render_html(realized: list[RealizedPnL], opt_events: list[OptionEvent],
                seed_lots: list[tuple[date, PositionLot, str]],
                year_range: tuple[int, int], file_list: list[str],
                leftover_lots: list, warnings: list[str],
                alias_map: dict, alias_unresolved: list[str],
                other_unhandled: list, skipped_pre_years: list[int],
                annual_summary: list[dict] | None = None) -> str:
    total_by_ccy: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    yearly_by_ccy: dict[tuple[int, str], Decimal] = defaultdict(lambda: Decimal("0"))
    monthly_by_ccy: dict[tuple[str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    by_year_asset: dict[tuple[int, str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    by_year_asset_cnt: dict[tuple[int, str, str], int] = defaultdict(int)
    by_subcat: dict[tuple[str, str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    by_subcat_cnt: dict[tuple[str, str, str], int] = defaultdict(int)
    by_underlying: dict[tuple[str, str, str], Decimal] = defaultdict(lambda: Decimal("0"))  # (ccy, underlying, asset_type)
    by_underlying_cnt: dict[tuple[str, str, str], int] = defaultdict(int)
    years_seen: set[int] = set()
    for r in realized:
        total_by_ccy[r.currency] += r.pnl
        yearly_by_ccy[(r.realized_year, r.currency)] += r.pnl
        ym = f"{r.close_date.year}-{r.close_date.month:02d}"
        monthly_by_ccy[(ym, r.currency)] += r.pnl
        by_year_asset[(r.realized_year, r.currency, r.asset_type)] += r.pnl
        by_year_asset_cnt[(r.realized_year, r.currency, r.asset_type)] += 1
        by_subcat[(r.realized_year, r.currency, r.asset_type, r.sub_category or "其他")] += r.pnl
        by_subcat_cnt[(r.realized_year, r.currency, r.asset_type, r.sub_category or "其他")] += 1
        und = r.underlying or r.code
        by_underlying[(r.currency, und, r.asset_type)] += r.pnl
        by_underlying_cnt[(r.currency, und, r.asset_type)] += 1
        years_seen.add(r.realized_year)
    currencies = sorted(total_by_ccy.keys())

    parts: list[str] = []
    parts.append("""<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="utf-8">
<title>已实现资本利得报表</title>
<style>
body{font-family:-apple-system,Segoe UI,Helvetica,Arial,sans-serif;margin:24px;color:#222;}
h1,h2{font-weight:600;}
.summary{display:flex;flex-wrap:wrap;gap:12px;margin:16px 0;}
.card{background:#f5f7fa;border:1px solid #d8dee9;border-radius:8px;padding:12px 16px;min-width:180px;}
.card .label{font-size:12px;color:#555;}
.card .value{font-size:20px;font-weight:600;margin-top:4px;}
.value.neg{color:#c0392b;} .value.pos{color:#1d8348;}
table{border-collapse:collapse;width:100%;margin:12px 0;font-size:13px;}
th,td{border:1px solid #d8dee9;padding:6px 8px;text-align:right;}
th{background:#eef1f5;}
td.code,td.note,td.date,td.left{text-align:left;}
tr.pos td.pnl{color:#1d8348;} tr.neg td.pnl{color:#c0392b;}
.toolbar{margin:8px 0;}
input[type=search],select{padding:4px 8px;border:1px solid #c0c8d2;border-radius:4px;}
details{margin:12px 0;}
summary{cursor:pointer;font-weight:600;font-size:16px;padding:6px 0;}
.meta{color:#666;font-size:12px;}
.warn{background:#fff7e6;border:1px solid #f0c97f;padding:8px 12px;border-radius:6px;margin:8px 0;}
.warn h3{margin:0 0 6px 0;font-size:14px;color:#a06a00;}
.warn ul{margin:4px 0;padding-left:20px;font-size:12px;color:#5f4500;}
</style></head><body>
""")
    parts.append(f"<h1>已实现资本利得报表 v2 · 移动加权平均 ({year_range[0]}–{year_range[1]})</h1>")
    parts.append("<div class='meta'>会计方法: <b>移动加权平均(MWA)</b> · "
                 "每只 (代码, 币种, 多/空) 维护单一 avg_cost; "
                 "开仓时按数量加权累加, 平仓时 avg_cost 不变。</div>")
    parts.append("<div class='meta'>数据源: " + ", ".join(_esc(Path(p).name) for p in file_list) + "</div>")
    if skipped_pre_years:
        parts.append(f"<div class='meta'>已跳过(无期初快照): {skipped_pre_years}</div>")

    parts.append("<div class='summary'>")
    for ccy in currencies:
        v = total_by_ccy[ccy]
        cls = "pos" if v >= 0 else "neg"
        parts.append(
            f"<div class='card'><div class='label'>{_esc(ccy)} 已实现总盈亏</div>"
            f"<div class='value {cls}'>{_fmt_dec(v)}</div></div>"
        )
    parts.append(
        f"<div class='card'><div class='label'>已实现笔数</div>"
        f"<div class='value'>{len(realized)}</div></div>"
        f"<div class='card'><div class='label'>期权特殊事件</div>"
        f"<div class='value'>{len(opt_events)}</div></div>"
        f"<div class='card'><div class='label'>种子持仓</div>"
        f"<div class='value'>{len(seed_lots)}</div></div>"
        f"<div class='card'><div class='label'>代码别名</div>"
        f"<div class='value'>{len(alias_map)}</div></div>"
    )
    parts.append("</div>")

    if warnings or alias_unresolved or other_unhandled:
        parts.append("<div class='warn'>")
        parts.append("<h3>⚠ 未处理事件清单 (人工核对建议)</h3>")
        if warnings:
            parts.append(f"<details open><summary>回放仓位/状态异常 ({len(warnings)})</summary><ul>")
            for w in warnings:
                parts.append(f"<li>{_esc(w)}</li>")
            parts.append("</ul></details>")
        if alias_unresolved:
            parts.append(f"<details><summary>未解析的 Option Adjustment ({len(alias_unresolved)})</summary><ul>")
            for w in alias_unresolved:
                parts.append(f"<li>{_esc(w)}</li>")
            parts.append("</ul></details>")
        if other_unhandled:
            parts.append(f"<details><summary>其他未处理资产进出 (DTC IN / Gift / 空备注等, {len(other_unhandled)})</summary>")
            parts.append("<ul>")
            for d, code, ccy, qty, note in other_unhandled:
                parts.append(f"<li>{_esc(d)} {_esc(code)} {_esc(ccy)} qty={_fmt_dec(qty, 4)} note={_esc(note)}</li>")
            parts.append("</ul></details>")
        parts.append("</div>")

    parts.append("<h2>年度小计 (按资产类型)</h2><table><tr><th>年份</th>")
    for ccy in currencies:
        parts.append(f"<th>{_esc(ccy)} 股票</th><th>{_esc(ccy)} 期权</th><th>{_esc(ccy)} 合计</th>")
    parts.append("</tr>")
    for yr in sorted(years_seen):
        parts.append(f"<tr><td class='date'>{yr}</td>")
        for ccy in currencies:
            stk = by_year_asset.get((yr, ccy, "股票"), Decimal("0"))
            opt = by_year_asset.get((yr, ccy, "期权"), Decimal("0"))
            tot = stk + opt
            for v in (stk, opt, tot):
                cls = "pos" if v >= 0 else "neg"
                parts.append(f"<td class='pnl {cls}'>{_fmt_dec(v)}</td>")
        parts.append("</tr>")
    # 总计行
    parts.append("<tr><td class='date'><b>合计</b></td>")
    for ccy in currencies:
        stk = sum((v for (y, c, a), v in by_year_asset.items() if c == ccy and a == "股票"), Decimal("0"))
        opt = sum((v for (y, c, a), v in by_year_asset.items() if c == ccy and a == "期权"), Decimal("0"))
        tot = stk + opt
        for v in (stk, opt, tot):
            cls = "pos" if v >= 0 else "neg"
            parts.append(f"<td class='pnl {cls}'><b>{_fmt_dec(v)}</b></td>")
    parts.append("</tr></table>")

    # ----- 年度交易概况 -----
    if annual_summary:
        parts.append("<h2>年度交易概况</h2>")
        parts.append("<table style='overflow-x:auto;display:block'><tr>"
                     "<th>年份</th><th>币种</th>"
                     "<th>年初持仓数</th><th>年初市值</th>"
                     "<th>买入笔数</th><th>买入金额</th>"
                     "<th>卖出笔数</th><th>卖出金额</th>"
                     "<th>已实现盈亏</th>"
                     "<th>年末持仓数</th><th>年末市值</th>"
                     "<th>浮动盈亏</th></tr>")
        for row in annual_summary:
            rpnl = Decimal(row["realized_pnl"])
            upnl = Decimal(row["unrealized_pnl"])
            rcls = "pos" if rpnl >= 0 else "neg"
            ucls = "pos" if upnl >= 0 else "neg"
            parts.append(
                f"<tr><td class='date'>{row['year']}</td>"
                f"<td>{_esc(row['currency'])}</td>"
                f"<td>{row['start_count']}</td>"
                f"<td>{_fmt_dec(Decimal(row['start_market_value']))}</td>"
                f"<td>{row['buy_count']}</td>"
                f"<td>{_fmt_dec(Decimal(row['buy_amount']))}</td>"
                f"<td>{row['sell_count']}</td>"
                f"<td>{_fmt_dec(Decimal(row['sell_amount']))}</td>"
                f"<td class='pnl {rcls}'>{_fmt_dec(rpnl)}</td>"
                f"<td>{row['end_count']}</td>"
                f"<td>{_fmt_dec(Decimal(row['end_market_value']))}</td>"
                f"<td class='pnl {ucls}'>{_fmt_dec(upnl)}</td></tr>"
            )
        # 合计行
        total_buy = sum((Decimal(r["buy_amount"]) for r in annual_summary), Decimal("0"))
        total_sell = sum((Decimal(r["sell_amount"]) for r in annual_summary), Decimal("0"))
        total_rpnl = sum((Decimal(r["realized_pnl"]) for r in annual_summary), Decimal("0"))
        total_upnl = sum((Decimal(r["unrealized_pnl"]) for r in annual_summary), Decimal("0"))
        total_buy_cnt = sum(r["buy_count"] for r in annual_summary)
        total_sell_cnt = sum(r["sell_count"] for r in annual_summary)
        rcls = "pos" if total_rpnl >= 0 else "neg"
        ucls = "pos" if total_upnl >= 0 else "neg"
        parts.append(
            f"<tr><td class='date'><b>合计</b></td><td></td>"
            f"<td></td><td></td>"
            f"<td><b>{total_buy_cnt}</b></td>"
            f"<td><b>{_fmt_dec(total_buy)}</b></td>"
            f"<td><b>{total_sell_cnt}</b></td>"
            f"<td><b>{_fmt_dec(total_sell)}</b></td>"
            f"<td class='pnl {rcls}'><b>{_fmt_dec(total_rpnl)}</b></td>"
            f"<td></td><td></td>"
            f"<td class='pnl {ucls}'><b>{_fmt_dec(total_upnl)}</b></td></tr>"
        )
        parts.append("</table>")

    # ----- 年度缴税估算 -----
    parts.append("<h2>年度缴税估算 (按 20% 资本利得税)</h2>")
    parts.append("<p style='color:#666;font-size:0.9em'>"
                 "口径: 按 (年份, 币种) 聚合, 当年某币种合计 PnL&gt;0 才计税(亏损不抵扣); "
                 f"USD/CNY 用年末中间价, HKD/CNY = USD/CNY ÷ {TAX_HKD_PEG_DIVISOR} (固定挂钩). "
                 "仅供参考, 实际申报口径以税务机关为准.</p>")
    parts.append("<table><tr><th>年份</th><th>币种</th><th>年度 PnL</th>"
                 "<th>应税额(20%)</th><th>年末汇率</th><th>折合 CNY 税额</th></tr>")
    tax_rows: list[tuple[int, str, Decimal, Decimal, Decimal, Decimal]] = []
    tax_total_cny = Decimal("0")
    for yr in sorted(years_seen):
        for ccy in currencies:
            pnl = sum((v for (y, c, a), v in by_year_asset.items() if y == yr and c == ccy),
                      Decimal("0"))
            usd_fx = TAX_FX_USD_TO_CNY.get(yr)
            if usd_fx is None:
                fx = None
            elif ccy == "USD":
                fx = usd_fx
            elif ccy == "HKD":
                fx = usd_fx / TAX_HKD_PEG_DIVISOR
            else:
                fx = None
            taxable = pnl if pnl > 0 else Decimal("0")
            tax_native = (taxable * TAX_RATE).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            tax_cny = (tax_native * fx).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) \
                if fx is not None else None
            tax_rows.append((yr, ccy, pnl, tax_native, fx, tax_cny))
            if tax_cny is not None:
                tax_total_cny += tax_cny
    for yr, ccy, pnl, tax_native, fx, tax_cny in tax_rows:
        cls = "pos" if pnl >= 0 else "neg"
        parts.append(
            f"<tr><td class='date'>{yr}</td><td>{_esc(ccy)}</td>"
            f"<td class='pnl {cls}'>{_fmt_dec(pnl)}</td>"
            f"<td class='pnl'>{_fmt_dec(tax_native) if tax_native > 0 else '—'}</td>"
            f"<td>{_fmt_dec(fx, 4) if fx is not None else '—'}</td>"
            f"<td class='pnl'>{_fmt_dec(tax_cny) if tax_cny is not None and tax_cny > 0 else '—'}</td></tr>"
        )
    parts.append(
        f"<tr><td colspan='5' class='date'><b>合计 CNY 税额</b></td>"
        f"<td class='pnl pos'><b>{_fmt_dec(tax_total_cny)}</b></td></tr>"
    )
    parts.append("</table>")

    parts.append("<h2>盈亏细分类别</h2>")
    sub_ccys = sorted({k[1] for k in by_subcat})
    sub_years = sorted({k[0] for k in by_subcat}, reverse=True)
    parts.append("<div class='toolbar'>年份 <select id='subYear'><option value=''>全部</option>")
    for yr in sub_years:
        parts.append(f"<option value='{yr}'>{yr}</option>")
    parts.append("</select> 币种 <select id='subCcy'><option value=''>全部</option>")
    for ccy in sub_ccys:
        parts.append(f"<option value='{_esc(ccy)}'>{_esc(ccy)}</option>")
    parts.append("</select></div>")
    parts.append("<table id='subTbl'><thead><tr>"
                 "<th>年</th><th>币种</th><th>资产</th><th>细分</th><th>笔数</th>"
                 "<th>实现盈亏</th></tr></thead><tbody>")
    for k in sorted(by_subcat):
        v = by_subcat[k]
        cls = "pos" if v >= 0 else "neg"
        parts.append(
            f"<tr data-year='{k[0]}' data-ccy='{_esc(k[1])}' "
            f"data-cnt='{by_subcat_cnt[k]}' data-pnl='{v}'>"
            f"<td>{k[0]}</td><td>{_esc(k[1])}</td><td>{_esc(k[2])}</td>"
            f"<td class='left'>{_esc(k[3])}</td>"
            f"<td>{by_subcat_cnt[k]}</td>"
            f"<td class='pnl {cls}'>{_fmt_dec(v)}</td></tr>"
        )
    parts.append("</tbody>"
                 "<tfoot><tr id='subSummary'><td colspan='6' class='left'>—</td></tr></tfoot>"
                 "</table>")

    parts.append("<h2>按底层资产汇总 (股票 + 期权合并)</h2>")
    # 聚合到 (ccy, underlying)
    by_und_total: dict[tuple[str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    by_und_stk: dict[tuple[str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    by_und_opt: dict[tuple[str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    for (ccy, und, atype), v in by_underlying.items():
        by_und_total[(ccy, und)] += v
        if atype == "股票":
            by_und_stk[(ccy, und)] += v
        else:
            by_und_opt[(ccy, und)] += v
    ranked = sorted(by_und_total.items(), key=lambda x: -abs(x[1]))
    parts.append("<table><tr><th>币种</th><th>底层</th><th>股票 PnL</th><th>期权 PnL</th><th>合计</th></tr>")
    for (ccy, und), v in ranked:
        stk = by_und_stk.get((ccy, und), Decimal("0"))
        opt = by_und_opt.get((ccy, und), Decimal("0"))
        cls_s = "pos" if stk >= 0 else "neg"
        cls_o = "pos" if opt >= 0 else "neg"
        cls_t = "pos" if v >= 0 else "neg"
        parts.append(
            f"<tr><td>{_esc(ccy)}</td><td class='code'>{_esc(und)}</td>"
            f"<td class='pnl {cls_s}'>{_fmt_dec(stk)}</td>"
            f"<td class='pnl {cls_o}'>{_fmt_dec(opt)}</td>"
            f"<td class='pnl {cls_t}'>{_fmt_dec(v)}</td></tr>"
        )
    parts.append("</table>")

    parts.append("<h2>月度盈亏汇总</h2><table><tr><th>月份</th>")
    for ccy in currencies:
        parts.append(f"<th>{_esc(ccy)}</th>")
    parts.append("</tr>")
    months = sorted({k[0] for k in monthly_by_ccy.keys()})
    for ym in months:
        parts.append(f"<tr><td class='date'>{ym}</td>")
        for ccy in currencies:
            v = monthly_by_ccy.get((ym, ccy), Decimal("0"))
            cls = "pos" if v >= 0 else "neg"
            parts.append(f"<td class='pnl {cls}'>{_fmt_dec(v)}</td>")
        parts.append("</tr>")
    parts.append("</table>")

    parts.append("<h2>已实现盈亏明细</h2>")
    subs_seen = sorted({r.sub_category for r in realized if r.sub_category})
    ccys_seen = sorted({r.currency for r in realized if r.currency})
    parts.append("<div class='toolbar'>"
                 "代码过滤 <input type='search' id='filterInput' placeholder='代码或底层'> "
                 "年份 <select id='filterYear'><option value=''>全部</option>")
    for yr in sorted(years_seen, reverse=True):
        parts.append(f"<option value='{yr}'>{yr}</option>")
    parts.append("</select> "
                 "资产 <select id='filterAsset'><option value=''>全部</option>"
                 "<option value='股票'>股票</option>"
                 "<option value='期权'>期权</option></select> "
                 "细分 <select id='filterSub'><option value=''>全部</option>")
    for sc in subs_seen:
        parts.append(f"<option value='{_esc(sc)}'>{_esc(sc)}</option>")
    parts.append("</select> 币种 <select id='filterCcy'><option value=''>全部</option>")
    for ccy in ccys_seen:
        parts.append(f"<option value='{_esc(ccy)}'>{_esc(ccy)}</option>")
    parts.append("</select></div>")

    parts.append("<table id='realTbl'><thead><tr>"
                 "<th>年</th><th>代码</th><th>底层</th><th>类型</th><th>细分</th><th>方向</th>"
                 "<th>数量</th><th>建仓日</th><th>建仓价</th><th>平仓日</th><th>平仓价</th>"
                 "<th>建仓费用</th><th>平仓费用</th><th>乘数</th><th>币种</th>"
                 "<th>实现盈亏</th><th>备注</th></tr></thead><tbody>")
    for r in sorted(realized, key=lambda x: x.close_date, reverse=True):
        cls = "pos" if r.pnl >= 0 else "neg"
        parts.append(
            f"<tr class='{cls}' data-code='{_esc(r.code)}' data-year='{r.realized_year}' "
            f"data-asset='{_esc(r.asset_type)}' data-sub='{_esc(r.sub_category)}' "
            f"data-ccy='{_esc(r.currency)}' data-pnl='{r.pnl}'>"
            f"<td>{r.realized_year}</td>"
            f"<td class='code'>{_esc(r.code)}</td>"
            f"<td class='code'>{_esc(r.underlying or r.code)}</td>"
            f"<td>{_esc(r.asset_type)}</td>"
            f"<td class='left'>{_esc(r.sub_category)}</td>"
            f"<td>{_esc(r.direction)}</td>"
            f"<td>{_fmt_dec(r.quantity, 4)}</td>"
            f"<td class='date'>{r.open_date}</td>"
            f"<td>{_fmt_dec(r.open_price, 4)}</td>"
            f"<td class='date'>{r.close_date}</td>"
            f"<td>{_fmt_dec(r.close_price, 4)}</td>"
            f"<td>{_fmt_dec(r.open_fee, 4)}</td>"
            f"<td>{_fmt_dec(r.close_fee, 4)}</td>"
            f"<td>{_fmt_dec(r.multiplier, 0)}</td>"
            f"<td>{_esc(r.currency)}</td>"
            f"<td class='pnl'>{_fmt_dec(r.pnl)}</td>"
            f"<td class='note'>{_esc(r.note)}</td></tr>"
        )
    parts.append("</tbody>"
                 "<tfoot><tr id='realSummary'><td colspan='17' class='left'>"
                 "—</td></tr></tfoot></table>")

    parts.append("<h2>期权指派/行权事件追踪</h2><table><tr>"
                 "<th>期权代码</th><th>事件</th><th>日期</th><th>合约数</th>"
                 "<th>行权价</th><th>权利金净额</th><th>已转股票?</th><th>备注</th></tr>")
    for e in sorted(opt_events, key=lambda x: x.event_date, reverse=True):
        parts.append(
            f"<tr><td class='code'>{_esc(e.code)}</td>"
            f"<td>{_esc(e.event_type)}</td>"
            f"<td class='date'>{e.event_date}</td>"
            f"<td>{_fmt_dec(e.contracts, 0)}</td>"
            f"<td>{'' if e.strike is None else _fmt_dec(e.strike, 3)}</td>"
            f"<td>{'' if e.premium_net is None else _fmt_dec(e.premium_net)}</td>"
            f"<td>{'是' if e.transferred_to_stock else '否'}</td>"
            f"<td class='note'>{_esc(e.note)}</td></tr>"
        )
    parts.append("</table>")

    parts.append("<details><summary>剩余未平仓持仓 (共 "
                 f"{len(leftover_lots)} 笔)</summary>")
    parts.append("<table><tr><th>方向</th><th>代码</th><th>币种</th><th>数量</th>"
                 "<th>建仓日</th><th>建仓价</th><th>乘数</th><th>备注</th></tr>")
    for side, lot in leftover_lots:
        parts.append(
            f"<tr><td>{_esc('多头' if side=='long' else '空头')}</td>"
            f"<td class='code'>{_esc(lot.code)}</td>"
            f"<td>{_esc(lot.currency)}</td>"
            f"<td>{_fmt_dec(lot.quantity, 4)}</td>"
            f"<td class='date'>{lot.open_date}</td>"
            f"<td>{_fmt_dec(lot.open_price, 4)}</td>"
            f"<td>{_fmt_dec(lot.multiplier, 0)}</td>"
            f"<td class='note'>{_esc(lot.note)}</td></tr>"
        )
    parts.append("</table></details>")

    parts.append("""
<script>
const inp = document.getElementById('filterInput');
const yr  = document.getElementById('filterYear');
const at  = document.getElementById('filterAsset');
const sb  = document.getElementById('filterSub');
const cc  = document.getElementById('filterCcy');
const tbl = document.getElementById('realTbl');
const sumCell = document.querySelector('#realSummary td');
const fmt = new Intl.NumberFormat('en-US', {minimumFractionDigits: 2, maximumFractionDigits: 2});
function applyFilter() {
  const kw = inp.value.trim().toLowerCase();
  const y  = yr.value;
  const a  = at.value;
  const s  = sb.value;
  const c  = cc.value;
  let count = 0;
  const sums = {};
  for (const tr of tbl.tBodies[0].rows) {
    const code = (tr.dataset.code || '').toLowerCase();
    const und  = tr.cells[2] ? tr.cells[2].textContent.toLowerCase() : '';
    const okKw = !kw || code.includes(kw) || und.includes(kw);
    const okYr = !y || tr.dataset.year === y;
    const okAt = !a || tr.dataset.asset === a;
    const okSb = !s || tr.dataset.sub === s;
    const okCc = !c || tr.dataset.ccy === c;
    const show = okKw && okYr && okAt && okSb && okCc;
    tr.style.display = show ? '' : 'none';
    if (show) {
      count++;
      const ccy = tr.dataset.ccy || '';
      const pnl = parseFloat(tr.dataset.pnl || '0');
      sums[ccy] = (sums[ccy] || 0) + pnl;
    }
  }
  const parts = [`总笔数: ${count}`];
  for (const ccy of Object.keys(sums).sort()) {
    parts.push(`${ccy}: ${fmt.format(sums[ccy])}`);
  }
  sumCell.textContent = parts.join(' | ');
}
inp.addEventListener('input', applyFilter);
yr.addEventListener('change', applyFilter);
at.addEventListener('change', applyFilter);
sb.addEventListener('change', applyFilter);
cc.addEventListener('change', applyFilter);
applyFilter();

const subY = document.getElementById('subYear');
const subC = document.getElementById('subCcy');
const subTbl = document.getElementById('subTbl');
const subSum = document.querySelector('#subSummary td');
function applySubFilter() {
  const y = subY.value;
  const c = subC.value;
  let cnt = 0;
  const sums = {};
  for (const tr of subTbl.tBodies[0].rows) {
    const show = (!y || tr.dataset.year === y) && (!c || tr.dataset.ccy === c);
    tr.style.display = show ? '' : 'none';
    if (show) {
      cnt += parseInt(tr.dataset.cnt || '0', 10);
      const ccy = tr.dataset.ccy || '';
      const pnl = parseFloat(tr.dataset.pnl || '0');
      sums[ccy] = (sums[ccy] || 0) + pnl;
    }
  }
  const parts = [`合计笔数: ${cnt}`];
  for (const ccy of Object.keys(sums).sort()) {
    parts.push(`${ccy}: ${fmt.format(sums[ccy])}`);
  }
  subSum.textContent = parts.join(' | ');
}
subY.addEventListener('change', applySubFilter);
subC.addEventListener('change', applySubFilter);
applySubFilter();
</script>
</body></html>""")
    return "".join(parts)


# --------------------------------------------------------------------------- #
# JSON output + CLI
# --------------------------------------------------------------------------- #


def _to_jsonable(obj):
    if isinstance(obj, Decimal):
        return str(obj)
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    raise TypeError(f"unserializable: {type(obj)}")


def _build_tax_estimate(by_year_asset: dict) -> dict:
    """根据 by_year_asset 复算各 (year, ccy) 的税额, 返回结构化 dict 供 JSON 输出。"""
    agg: dict[tuple[int, str], Decimal] = defaultdict(lambda: Decimal("0"))
    for (y, c, _a), v in by_year_asset.items():
        agg[(y, c)] += v
    rows = []
    total_cny = Decimal("0")
    for (yr, ccy), pnl in sorted(agg.items()):
        usd_fx = TAX_FX_USD_TO_CNY.get(yr)
        if usd_fx is None:
            fx = None
        elif ccy == "USD":
            fx = usd_fx
        elif ccy == "HKD":
            fx = usd_fx / TAX_HKD_PEG_DIVISOR
        else:
            fx = None
        taxable = pnl if pnl > 0 else Decimal("0")
        tax_native = (taxable * TAX_RATE).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        tax_cny = (tax_native * fx).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) \
            if fx is not None else None
        if tax_cny is not None:
            total_cny += tax_cny
        rows.append({
            "year": yr, "currency": ccy,
            "pnl": str(pnl), "tax_native": str(tax_native),
            "fx_to_cny": str(fx) if fx is not None else None,
            "tax_cny": str(tax_cny) if tax_cny is not None else None,
        })
    return {"tax_rate": str(TAX_RATE), "rows": rows, "total_cny": str(total_cny)}


def render_json(realized, opt_events, seed_lots, leftover_lots,
                year_range, file_list, warnings, alias_map,
                alias_unresolved, other_unhandled,
                annual_summary=None) -> str:
    # 统计分类汇总,放进 summary 方便外部消费
    by_year_asset: dict[tuple, Decimal] = defaultdict(lambda: Decimal("0"))
    by_subcat: dict[tuple, Decimal] = defaultdict(lambda: Decimal("0"))
    by_underlying: dict[tuple, Decimal] = defaultdict(lambda: Decimal("0"))
    for r in realized:
        by_year_asset[(r.realized_year, r.currency, r.asset_type)] += r.pnl
        by_subcat[(r.realized_year, r.currency, r.asset_type, r.sub_category)] += r.pnl
        by_underlying[(r.currency, r.underlying or r.code, r.asset_type)] += r.pnl
    payload = {
        "year_range": list(year_range),
        "files": file_list,
        "summary": {
            "realized_count": len(realized),
            "option_events_count": len(opt_events),
            "seed_lots_count": len(seed_lots),
            "leftover_lots_count": len(leftover_lots),
            "warnings_count": len(warnings),
            "alias_map_count": len(alias_map),
            "other_unhandled_count": len(other_unhandled),
            "by_year_asset": [
                {"year": y, "currency": c, "asset_type": a, "pnl": str(v)}
                for (y, c, a), v in sorted(by_year_asset.items())
            ],
            "by_sub_category": [
                {"year": y, "currency": c, "asset_type": a, "sub_category": s, "pnl": str(v)}
                for (y, c, a, s), v in sorted(by_subcat.items())
            ],
            "by_underlying": [
                {"currency": c, "underlying": u, "asset_type": a, "pnl": str(v)}
                for (c, u, a), v in sorted(by_underlying.items())
            ],
            "tax_estimate": _build_tax_estimate(by_year_asset),
        },
        "alias_map": alias_map,
        "alias_unresolved": alias_unresolved,
        "warnings": warnings,
        "other_unhandled": [
            {"date": d, "code": code, "currency": ccy, "quantity": qty, "note": note}
            for d, code, ccy, qty, note in other_unhandled
        ],
        "realized": [asdict(r) for r in realized],
        "option_events": [asdict(e) for e in opt_events],
        "leftover_lots": [
            {"side": side, **asdict(lot)} for side, lot in leftover_lots
        ],
        "annual_summary": annual_summary or [],
    }
    return json.dumps(payload, ensure_ascii=False, indent=2, default=_to_jsonable)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("target_dir", nargs="?", default=".",
                    help="Directory containing YYYY_年度账单_*.xlsx (default: current dir)")
    args = ap.parse_args()
    target = Path(args.target_dir).resolve()

    # 加载并校验配置文件
    cfg = load_config(target)
    apply_config(cfg)

    data = load_all_years(target)

    # cost_basis_overrides.json 可进一步覆盖 script_config.json 中的税务配置
    global TAX_FX_USD_TO_CNY, TAX_RATE
    tax_config = data.get("tax_config", {})
    if "tax_rate" in tax_config:
        TAX_RATE = tax_config["tax_rate"]
    if "fx_rates" in tax_config:
        TAX_FX_USD_TO_CNY.update(tax_config["fx_rates"])

    realized, opt_events, book, warnings, year_end_books = replay(
        data["trades"], data["movements"], data["seed_lots"]
    )
    # 把 cost_basis_overrides 的处理信息前置到 warnings 区,方便用户核对
    warnings = list(data.get("cost_basis_overrides", [])) + warnings
    leftover = list(book.iter_open())

    # 计算年度交易概况
    annual_summary = compute_annual_summary(
        data["trades"], realized,
        data.get("position_snapshots", []),
        year_end_books, data["year_range"],
    )

    html_str = render_html(
        realized, opt_events, data["seed_lots"],
        data["year_range"], data["files"], leftover, warnings,
        data.get("alias_map", {}), data.get("alias_unresolved", []),
        data.get("other_unhandled", []), data.get("skipped_pre_years", []),
        annual_summary,
    )
    json_str = render_json(
        realized, opt_events, data["seed_lots"], leftover,
        data["year_range"], data["files"], warnings,
        data.get("alias_map", {}), data.get("alias_unresolved", []),
        data.get("other_unhandled", []),
        annual_summary,
    )
    html_path = target / OUTPUT_HTML
    json_path = target / OUTPUT_JSON
    html_path.write_text(html_str, encoding="utf-8")
    json_path.write_text(json_str, encoding="utf-8")
    print("[v2 移动加权平均]")
    print(f"已生成: {html_path}")
    print(f"已生成: {json_path}")
    print(f"涉及年份: {data['year_range']}, 已实现 {len(realized)} 笔, "
          f"期权特殊事件 {len(opt_events)} 笔, 剩余未平仓 {len(leftover)} 笔, "
          f"warnings {len(warnings)} 条")
    if data.get("alias_map"):
        print(f"代码别名归一化: {len(data['alias_map'])} 项")
    if data.get("other_unhandled"):
        print(f"未处理资产进出: {len(data['other_unhandled'])} 条 (DTC IN / Gift / 空备注)")
    # 按 (币种, 年, 资产类型) 拆分
    by_yat: dict[tuple[str, int, str], Decimal] = defaultdict(lambda: Decimal("0"))
    by_yat_cnt: dict[tuple[str, int, str], int] = defaultdict(int)
    for r in realized:
        by_yat[(r.currency, r.realized_year, r.asset_type)] += r.pnl
        by_yat_cnt[(r.currency, r.realized_year, r.asset_type)] += 1
    print()
    print("=== 已实现盈亏(按 币种/年/资产类型) ===")
    for k in sorted(by_yat):
        print(f"  {k[0]} {k[1]} {k[2]:>4}: {by_yat[k]:>14,.2f}  ({by_yat_cnt[k]} 笔)")
    # 细分
    by_sub: dict[tuple[str, str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    by_sub_cnt: dict[tuple[str, str, str], int] = defaultdict(int)
    for r in realized:
        by_sub[(r.currency, r.asset_type, r.sub_category)] += r.pnl
        by_sub_cnt[(r.currency, r.asset_type, r.sub_category)] += 1
    print()
    print("=== 细分类别 ===")
    for k in sorted(by_sub):
        print(f"  {k[0]} {k[1]:>4} {k[2]:>10}: {by_sub[k]:>14,.2f}  ({by_sub_cnt[k]} 笔)")
    # 底层 Top
    by_und: dict[tuple[str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    for r in realized:
        by_und[(r.currency, r.underlying or r.code)] += r.pnl
    ranked = sorted(by_und.items(), key=lambda x: -abs(x[1]))[:10]
    print()
    print("=== 底层资产 Top 10 (股票+期权合并, 按绝对值排序) ===")
    for (k, v) in ranked:
        print(f"  {k[0]} {k[1]:>10}: {v:>14,.2f}")

    # 年度交易概况
    if annual_summary:
        def _f(v):
            return f"{Decimal(v):>14,.2f}" if isinstance(v, str) else f"{v:>14,.2f}"
        print()
        print("=== 年度交易概况 ===")
        for row in annual_summary:
            print(
                f"  {row['year']} {row['currency']}: "
                f"年初{row['start_count']}只/{_f(row['start_market_value'])}, "
                f"买入{row['buy_count']}笔/{_f(row['buy_amount'])}, "
                f"卖出{row['sell_count']}笔/{_f(row['sell_amount'])}, "
                f"已实现{_f(row['realized_pnl'])}, "
                f"年末{row['end_count']}只/{_f(row['end_market_value'])}, "
                f"浮动{_f(row['unrealized_pnl'])}"
            )


if __name__ == "__main__":
    main()
